#!/usr/bin/env python3
"""
Script to check Excel formatting in the output file
"""
import openpyxl
import sys
from pathlib import Path

def check_excel_formatting(file_path):
    """Check the formatting of an Excel file"""
    try:
        print(f"Checking file: {file_path}")
        
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)
        
        print(f"Available sheets: {wb.sheetnames}")
        
        if 'RBMF' in wb.sheetnames:
            ws = wb['RBMF']
            print(f"\n=== RBMF Tab Info ===")
            print(f"Total rows: {ws.max_row}")
            print(f"Total columns: {ws.max_column}")
            
            # Check header row formatting
            print(f"\n=== Header Row (Row 1) Formatting ===")
            for col in range(1, min(6, ws.max_column + 1)):
                cell = ws.cell(row=1, column=col)
                font_size = cell.font.size if cell.font else "None"
                wrap_text = cell.alignment.wrap_text if cell.alignment else "None"
                print(f"Column {col}: Font size = {font_size}, Wrap text = {wrap_text}")
            
            # Check column widths
            print(f"\n=== Column Widths ===")
            for col in range(1, min(6, ws.max_column + 1)):
                col_letter = openpyxl.utils.get_column_letter(col)
                width = ws.column_dimensions[col_letter].width
                print(f"Column {col_letter}: {width}")
            
            # Check row heights
            print(f"\n=== Row Heights (first 5 rows) ===")
            for row in range(1, min(6, ws.max_row + 1)):
                height = ws.row_dimensions[row].height
                print(f"Row {row}: {height}")
            
            # Check data cell formatting
            print(f"\n=== Data Cell Formatting (Row 2) ===")
            for col in range(1, min(6, ws.max_column + 1)):
                cell = ws.cell(row=2, column=col)
                font_size = cell.font.size if cell.font else "None"
                wrap_text = cell.alignment.wrap_text if cell.alignment else "None"
                print(f"Column {col}: Font size = {font_size}, Wrap text = {wrap_text}")
                
        else:
            print("RBMF tab not found!")
            
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    file_path = "data/2025-output/test/final/INO_2024_Aquatera_Energy Transition Business and Change Management Centre of Excellence.xlsx"
    
    if not Path(file_path).exists():
        print(f"File not found: {file_path}")
        sys.exit(1)
    
    success = check_excel_formatting(file_path)
    sys.exit(0 if success else 1)
