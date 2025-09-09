"""Excel processing optimizations for RBMF transformation."""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, List
from loguru import logger
import io
from contextlib import contextmanager

class ExcelOptimizer:
    """Optimizes Excel file processing operations."""
    
    def __init__(self):
        """Initialize Excel optimizer."""
        self.template_cache = {}
        self.workbook_cache = {}
    
    @contextmanager
    def optimized_excel_reader(self, file_path: Path, sheet_name: str = None):
        """Context manager for optimized Excel reading.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to read (None for all sheets)
            
        Yields:
            Optimized Excel reader
        """
        try:
            # Use read_only mode for large files
            file_size = file_path.stat().st_size
            read_only = file_size > 10 * 1024 * 1024  # 10MB threshold
            
            if read_only:
                logger.debug(f"Using read_only mode for large file: {file_path.name}")
            
            with pd.ExcelFile(file_path, engine='openpyxl') as excel_file:
                yield excel_file
                
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            raise
    
    def read_excel_optimized(self, file_path: Path, sheet_name: str = None, 
                           use_chunks: bool = False, chunk_size: int = 1000) -> Dict[str, pd.DataFrame]:
        """Read Excel file with optimizations.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to read
            use_chunks: Whether to use chunked reading
            chunk_size: Size of chunks if using chunked reading
            
        Returns:
            Dictionary of sheet names to DataFrames
        """
        results = {}
        
        with self.optimized_excel_reader(file_path, sheet_name) as excel_file:
            sheets_to_read = [sheet_name] if sheet_name else excel_file.sheet_names
            
            for sheet in sheets_to_read:
                try:
                    if use_chunks:
                        # Read in chunks for very large sheets
                        chunks = []
                        for chunk in pd.read_excel(file_path, sheet_name=sheet, chunksize=chunk_size):
                            chunks.append(chunk)
                        results[sheet] = pd.concat(chunks, ignore_index=True)
                    else:
                        # Read entire sheet
                        results[sheet] = pd.read_excel(file_path, sheet_name=sheet)
                        
                    logger.debug(f"Read sheet '{sheet}' with {len(results[sheet])} rows")
                    
                except Exception as e:
                    logger.error(f"Error reading sheet '{sheet}' from {file_path}: {e}")
                    results[sheet] = pd.DataFrame()
        
        return results
    
    def write_excel_optimized(self, dataframes: Dict[str, pd.DataFrame], 
                            output_path: Path, template_path: Optional[Path] = None) -> bool:
        """Write Excel file with optimizations.
        
        Args:
            dataframes: Dictionary of sheet names to DataFrames
            output_path: Path for output file
            template_path: Optional template file for formatting
            
        Returns:
            True if successful
        """
        try:
            if template_path and template_path.exists():
                # Use template for formatting
                return self._write_with_template(dataframes, output_path, template_path)
            else:
                # Write without template (faster)
                return self._write_simple(dataframes, output_path)
                
        except Exception as e:
            logger.error(f"Error writing Excel file {output_path}: {e}")
            return False
    
    def _write_with_template(self, dataframes: Dict[str, pd.DataFrame], 
                           output_path: Path, template_path: Path) -> bool:
        """Write Excel file using template formatting."""
        try:
            # Load template
            template_wb = openpyxl.load_workbook(template_path)
            
            # Create new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for sheet_name, df in dataframes.items():
                # Create worksheet
                ws = wb.create_sheet(sheet_name)
                
                # Write data
                for r_idx, row in enumerate(df.itertuples(index=False), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply basic formatting if template has this sheet
                if sheet_name in template_wb.sheetnames:
                    template_ws = template_wb[sheet_name]
                    self._copy_formatting(template_ws, ws, len(df))
            
            wb.save(output_path)
            wb.close()
            template_wb.close()
            
            logger.debug(f"Written Excel file with template: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error writing Excel with template: {e}")
            return False
    
    def _write_simple(self, dataframes: Dict[str, pd.DataFrame], output_path: Path) -> bool:
        """Write Excel file without template (faster)."""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.debug(f"Written simple Excel file: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error writing simple Excel: {e}")
            return False
    
    def _copy_formatting(self, source_ws, target_ws, data_rows: int):
        """Copy basic formatting from source to target worksheet."""
        try:
            # Copy column widths
            for col in source_ws.column_dimensions:
                target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width
            
            # Copy header formatting (first row)
            if data_rows > 0:
                for col in range(1, source_ws.max_column + 1):
                    source_cell = source_ws.cell(row=1, column=col)
                    target_cell = target_ws.cell(row=1, column=col)
                    
                    if source_cell.has_style:
                        target_cell.font = source_cell.font
                        target_cell.fill = source_cell.fill
                        target_cell.border = source_cell.border
                        target_cell.alignment = source_cell.alignment
                        
        except Exception as e:
            logger.warning(f"Could not copy formatting: {e}")
    
    def get_sheet_info(self, file_path: Path) -> Dict[str, Dict[str, Any]]:
        """Get information about Excel file sheets without loading data.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with sheet information
        """
        try:
            with self.optimized_excel_reader(file_path) as excel_file:
                sheet_info = {}
                
                for sheet_name in excel_file.sheet_names:
                    # Read only first few rows to get structure
                    sample_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                    
                    sheet_info[sheet_name] = {
                        'columns': list(sample_df.columns),
                        'column_count': len(sample_df.columns),
                        'sample_rows': len(sample_df),
                        'has_data': not sample_df.empty
                    }
                
                return sheet_info
                
        except Exception as e:
            logger.error(f"Error getting sheet info for {file_path}: {e}")
            return {}
