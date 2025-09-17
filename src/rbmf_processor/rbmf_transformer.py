"""RBMF Data Transformer for creating files with Instructions tab from template."""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from copy import copy
from loguru import logger
import json

from .config import settings


class RBMFTransformer:
    """Creates files with Instructions tab from template."""
    
    def __init__(self, data_dir: Path, include_steps: bool = False, target_folders: list = None):
        """Initialize the RBMF transformer.
        
        Args:
            data_dir: Path to the data directory containing source files and template
            include_steps: If True, create intermediate tabs (RBMF_1, RBMF_2) for verification
            target_folders: List of specific folders to process. If None, will process all available folders.
        """
        self.data_dir = Path(data_dir)
        self.template_file = self.data_dir / "Draft RBMF Template 2025.xlsx"
        self.output_dir = self.data_dir / "2025-output"
        self.include_steps = include_steps
        self.template_instructions = None
        self.file_to_project_id_mapping = None
        
        # Set target folders
        if target_folders is None:
            # Default behavior: process all available folders
            self.target_folders = self.discover_available_folders()
        else:
            self.target_folders = target_folders
    
    def load_file_to_project_id_mapping(self) -> Dict[str, str]:
        """Load the file_to_projectId_mapping.json file.
        
        Returns:
            Dictionary mapping file names to Project IDs
            
        Raises:
            FileNotFoundError: If mapping file doesn't exist
            ValueError: If mapping file is invalid
        """
        mapping_file = self.output_dir / "file_to_projectId_mapping.json"
        
        if not mapping_file.exists():
            logger.warning(f"Mapping file not found: {mapping_file}")
            return {}
        
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                raw_data = json.load(f)
            
            # Extract mappings from nested structure
            self.file_to_project_id_mapping = {}
            
            if 'folders' in raw_data:
                for folder_name, folder_data in raw_data['folders'].items():
                    if 'mappings' in folder_data:
                        for file_name, project_id in folder_data['mappings'].items():
                            self.file_to_project_id_mapping[file_name] = project_id
                            logger.debug(f"Mapped {file_name} -> {project_id}")
            
            logger.info(f"Loaded file-to-project mapping with {len(self.file_to_project_id_mapping)} entries")
            return self.file_to_project_id_mapping
            
        except Exception as e:
            logger.error(f"Error loading file-to-project mapping: {e}")
            raise ValueError(f"Could not load file-to-project mapping: {e}")
    
    def get_project_id_for_file(self, file_name: str) -> str:
        """Get Project ID for a given file name.
        
        Args:
            file_name: Base name of the file (without path)
            
        Returns:
            Project ID if found, "NOT FOUND" otherwise
        """
        if self.file_to_project_id_mapping is None:
            logger.warning("File-to-project mapping not loaded, returning 'NOT FOUND'")
            return "NOT FOUND"
        
        # Use base name for lookup
        base_name = Path(file_name).name
        
        project_id = self.file_to_project_id_mapping.get(base_name, "NOT FOUND")
        
        if project_id == "NOT FOUND":
            logger.warning(f"Project ID not found for file: {base_name}")
        else:
            logger.info(f"Found Project ID '{project_id}' for file: {base_name}")
        
        return project_id
    
    def load_column_mapping(self) -> List[Dict[str, str]]:
        """Load column value mapping from JSON file.
        
        Returns:
            List of mapping dictionaries with column, original_value, new_value
        """
        mapping_file = self.output_dir / "column_mapping.json"
        
        if not mapping_file.exists():
            logger.warning(f"Column mapping not found: {mapping_file}")
            return []
        
        try:
            with open(mapping_file, 'r', encoding='utf-8') as f:
                mappings = json.load(f)
                logger.info(f"Loaded column mapping with {len(mappings)} entries")
                return mappings
        except Exception as e:
            logger.error(f"Error loading column mapping: {e}")
            return []
    
    def _fuzzy_match_value(self, value: str, mappings: List[Dict[str, str]], column_name: str, threshold: int = 90) -> str:
        """Find the best fuzzy match for a value in the mappings.
        
        Args:
            value: The value to match
            column_name: The column name to filter mappings
            mappings: List of mapping dictionaries
            threshold: Fuzzy matching threshold (0-100)
            
        Returns:
            The new value if match found, original value otherwise
        """
        if not value or not isinstance(value, str):
            return value
        
        # Filter mappings for the specific column
        column_mappings = [m for m in mappings if m.get('column') == column_name]
        
        if not column_mappings:
            return value
        
        best_match = None
        best_score = 0
        
        for mapping in column_mappings:
            original = mapping.get('original_value', '')
            if not original:
                continue
                
            # Calculate fuzzy match score
            from fuzzywuzzy import fuzz
            score = fuzz.ratio(value.strip(), original.strip())
            
            if score >= threshold and score > best_score:
                best_match = mapping
                best_score = score
        
        if best_match:
            new_value = best_match.get('new_value', '')
            if new_value:  # Only replace if new_value is not empty
                logger.info(f"Fuzzy matched '{value}' -> '{new_value}' (score: {best_score}%)")
                return new_value
            else:
                logger.warning(f"Found match for '{value}' but new_value is empty, keeping original")
                return value
        
        logger.debug(f"No fuzzy match found for '{value}' in column '{column_name}' (threshold: {threshold}%)")
        return value
    
    def _find_best_indicator_mapping(self, indicator_name: str, mappings: List[Dict[str, str]], threshold: int = 90) -> Optional[Dict[str, str]]:
        """Find the best fuzzy match for an indicator name in the mappings.
        
        Args:
            indicator_name: The indicator name to match
            mappings: List of mapping dictionaries
            threshold: Fuzzy matching threshold (0-100)
            
        Returns:
            The best matching mapping dictionary or None if no match found
        """
        if not indicator_name or not isinstance(indicator_name, str):
            return None
        
        # Filter mappings for Indicator name column
        indicator_mappings = [m for m in mappings if m.get('column') == 'Indicator name']
        
        if not indicator_mappings:
            return None
        
        best_match = None
        best_score = 0
        
        for mapping in indicator_mappings:
            original = mapping.get('original_value', '')
            if not original:
                continue
                
            # Calculate fuzzy match score
            from fuzzywuzzy import fuzz
            score = fuzz.ratio(indicator_name.strip(), original.strip())
            
            if score >= threshold and score > best_score:
                best_match = mapping
                best_score = score
        
        if best_match:
            logger.info(f"Found indicator mapping: '{indicator_name}' -> '{best_match.get('new_value', '')}' (score: {best_score}%)")
        
        return best_match
    
    def _apply_column_mapping(self, df: pd.DataFrame, mappings: List[Dict[str, str]]) -> pd.DataFrame:
        """Apply column value mappings to a DataFrame.
        
        Args:
            df: The DataFrame to modify
            mappings: List of mapping dictionaries
            
        Returns:
            Modified DataFrame with mapped values
        """
        if not mappings:
            return df
        
        df_mapped = df.copy()
        
        # Apply mappings to Strategic Outcome column
        if 'Strategic Outcome' in df_mapped.columns:
            logger.info("Applying mappings to Strategic Outcome column")
            df_mapped['Strategic Outcome'] = df_mapped['Strategic Outcome'].apply(
                lambda x: self._fuzzy_match_value(x, mappings, 'Strategic Outcome')
            )
        
        # Apply mappings to Indicator name and Indicator ID columns together
        if 'Indicator name' in df_mapped.columns:
            logger.info("Applying mappings to Indicator name and Indicator ID columns")
            
            # Process each row to update both Indicator name and Indicator ID
            for idx, row in df_mapped.iterrows():
                indicator_name = row['Indicator name']
                if indicator_name and isinstance(indicator_name, str):
                    # Find the best match for the indicator name
                    best_match = self._find_best_indicator_mapping(indicator_name, mappings)
                    
                    if best_match:
                        # Update Indicator name
                        df_mapped.at[idx, 'Indicator name'] = best_match['new_value']
                        
                        # Update Indicator ID if the column exists and we have a new indicator ID
                        if 'Indicator ID' in df_mapped.columns and best_match.get('new_indicator_id'):
                            df_mapped.at[idx, 'Indicator ID'] = best_match['new_indicator_id']
                            logger.info(f"Updated Indicator ID: {row.get('Indicator ID', 'N/A')} -> {best_match['new_indicator_id']}")
        
        return df_mapped
    
    def discover_available_folders(self) -> list:
        """Discover available folders in the data directory that contain Excel files.
        
        Returns:
            List of folder names that contain processable files
        """
        available_folders = []
        
        if not self.data_dir.exists():
            logger.warning(f"Data directory does not exist: {self.data_dir}")
            return available_folders
        
        # Folders to exclude (system folders, output folders, etc.)
        excluded_folders = {
            '2025-output', '2025-output-final', '.git', '__pycache__', 
            '.pytest_cache', 'node_modules', '.vscode', '.idea'
        }
        
        for item in self.data_dir.iterdir():
            if (item.is_dir() and 
                item.name not in excluded_folders and
                not item.name.startswith('.')):
                
                # Check if folder contains Excel files
                has_excel_files = False
                for file_path in item.glob('*'):
                    if (file_path.is_file() and 
                        (file_path.suffix.lower() in ['.xlsx', '.xls'] or 
                         self._is_excel_file(file_path))):
                        has_excel_files = True
                        break
                
                if has_excel_files:
                    available_folders.append(item.name)
        
        # Sort folders for consistent ordering
        available_folders.sort()
        
        if available_folders:
            logger.info(f"Discovered {len(available_folders)} folders with Excel files: {available_folders}")
        else:
            logger.warning("No folders with Excel files found in data directory")
        
        return available_folders
    
    def validate_folders(self, folder_names: list) -> tuple[list, list]:
        """Validate that specified folders exist and contain Excel files.
        
        Args:
            folder_names: List of folder names to validate
            
        Returns:
            Tuple of (valid_folders, invalid_folders)
        """
        valid_folders = []
        invalid_folders = []
        available_folders = self.discover_available_folders()
        
        for folder_name in folder_names:
            if folder_name in available_folders:
                valid_folders.append(folder_name)
            else:
                invalid_folders.append(folder_name)
        
        if invalid_folders:
            logger.warning(f"Invalid folders specified: {invalid_folders}")
            logger.info(f"Available folders: {available_folders}")
            
            # Suggest similar folder names for typos
            for invalid_folder in invalid_folders:
                suggestions = self._suggest_similar_folders(invalid_folder, available_folders)
                if suggestions:
                    logger.info(f"Did you mean one of these? {suggestions}")
        
        return valid_folders, invalid_folders
    
    def _suggest_similar_folders(self, target: str, available: list, max_suggestions: int = 3) -> list:
        """Suggest similar folder names for typos.
        
        Args:
            target: The invalid folder name
            available: List of available folder names
            max_suggestions: Maximum number of suggestions to return
            
        Returns:
            List of similar folder names
        """
        suggestions = []
        target_lower = target.lower()
        
        for folder in available:
            folder_lower = folder.lower()
            
            # Check for exact substring match
            if target_lower in folder_lower or folder_lower in target_lower:
                suggestions.append(folder)
            # Check for similar length and character overlap
            elif (abs(len(target) - len(folder)) <= 2 and 
                  len(set(target_lower) & set(folder_lower)) >= len(target_lower) * 0.6):
                suggestions.append(folder)
        
        return suggestions[:max_suggestions]
    
    def _validate_google_sheets_compatibility(self, file_path: Path) -> None:
        """Validate that the created file is compatible with Google Sheets.
        
        Args:
            file_path: Path to the Excel file to validate
        """
        try:
            # Check file size (Google Sheets limit: 10MB)
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            if file_size_mb > 10:
                logger.warning(f"File {file_path.name} is {file_size_mb:.2f}MB, which exceeds Google Sheets 10MB limit")
            elif file_size_mb > 5:
                logger.info(f"File {file_path.name} is {file_size_mb:.2f}MB, approaching Google Sheets limit")
            else:
                logger.info(f"File {file_path.name} is {file_size_mb:.2f}MB, within Google Sheets limits")
            
            # Check file extension
            if not file_path.suffix.lower() == '.xlsx':
                logger.warning(f"File {file_path.name} does not have .xlsx extension, may not be compatible with Google Sheets")
            
            # Check workbook structure
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                
                # Check number of sheets (Google Sheets limit: 200 sheets)
                if len(wb.sheetnames) > 200:
                    logger.warning(f"File {file_path.name} has {len(wb.sheetnames)} sheets, which exceeds Google Sheets 200 sheet limit")
                
                # Check each sheet for compatibility
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Check row count (Google Sheets limit: 5 million rows)
                    if ws.max_row > 5000000:
                        logger.warning(f"Sheet '{sheet_name}' in {file_path.name} has {ws.max_row} rows, which exceeds Google Sheets 5M row limit")
                    
                    # Check column count (Google Sheets limit: 18,278 columns)
                    if ws.max_column > 18278:
                        logger.warning(f"Sheet '{sheet_name}' in {file_path.name} has {ws.max_column} columns, which exceeds Google Sheets 18,278 column limit")
                    
                    # Check total cell count (Google Sheets limit: 5 million cells per sheet)
                    total_cells = ws.max_row * ws.max_column
                    if total_cells > 5000000:
                        logger.warning(f"Sheet '{sheet_name}' in {file_path.name} has {total_cells} cells, which exceeds Google Sheets 5M cell limit")
                
                wb.close()
                logger.info(f"Google Sheets compatibility validation completed for {file_path.name}")
                
            except Exception as e:
                logger.warning(f"Could not validate Google Sheets compatibility for {file_path.name}: {e}")
                
        except Exception as e:
            logger.warning(f"Error during Google Sheets compatibility validation for {file_path.name}: {e}")
        
    def load_template_instructions(self) -> pd.DataFrame:
        """Load the Instructions tab from the template file.
        
        Returns:
            DataFrame containing the Instructions tab data
            
        Raises:
            FileNotFoundError: If template file doesn't exist
            ValueError: If Instructions tab is not found
        """
        if not self.template_file.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_file}")
        
        try:
            # Read the Instructions tab
            instructions_df = pd.read_excel(self.template_file, sheet_name='Instructions')
            logger.info(f"Loaded Instructions template with {len(instructions_df)} rows")
            self.template_instructions = instructions_df
            return instructions_df
            
        except Exception as e:
            logger.error(f"Error loading template Instructions tab: {e}")
            raise ValueError(f"Could not load Instructions tab from template: {e}")
    
    def load_template_workbook(self) -> openpyxl.Workbook:
        """Load the template workbook with openpyxl to preserve formatting.
        
        Returns:
            openpyxl Workbook object
            
        Raises:
            FileNotFoundError: If template file doesn't exist
            ValueError: If Instructions tab is not found
        """
        if not self.template_file.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_file}")
        
        try:
            # Load the template workbook with openpyxl
            template_wb = openpyxl.load_workbook(self.template_file)
            
            # Check if Instructions tab exists
            if 'Instructions' not in template_wb.sheetnames:
                raise ValueError("Instructions tab not found in template file")
            
            logger.info(f"Loaded template workbook with sheets: {template_wb.sheetnames}")
            return template_wb
            
        except Exception as e:
            logger.error(f"Error loading template workbook: {e}")
            raise ValueError(f"Could not load template workbook: {e}")
    
    def create_output_structure(self, output_mode: str = None) -> None:
        """Create the 2025-output folder structure and clean existing files.
        
        Args:
            output_mode: Output mode ('steps', 'final', 'final-filter'). If None, uses include_steps flag.
        """
        self.output_dir.mkdir(exist_ok=True)
        
        for folder in self.target_folders:
            # Create subfolder based on mode
            if output_mode:
                folder_path = self.output_dir / folder / output_mode
            elif self.include_steps:
                folder_path = self.output_dir / folder / "steps"
            else:
                folder_path = self.output_dir / folder / "final"
            
            folder_path.mkdir(exist_ok=True)
            
            # Clean up existing files in the output folder
            if folder_path.exists():
                for file_path in folder_path.glob('*'):
                    if file_path.is_file():
                        try:
                            file_path.unlink()
                            logger.info(f"Deleted old file: {file_path.name}")
                        except Exception as e:
                            logger.warning(f"Could not delete {file_path.name}: {e}")
            
            mode_name = output_mode or ("steps" if self.include_steps else "final")
            logger.info(f"Created/cleaned {mode_name} output folder: {folder_path}")
    
    def create_instructions_only_file(self, output_file: Path) -> bool:
        """Create a file with only the Instructions tab from template.
        
        Args:
            output_file: Path for the output file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if self.template_instructions is None:
                raise ValueError("Template instructions not loaded. Call load_template_instructions() first.")
            
            # Load template workbook to preserve formatting
            template_wb = self.load_template_workbook()
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Copy Instructions tab with all formatting from template
            template_instructions_ws = template_wb['Instructions']
            ws_instructions = wb.create_sheet("Instructions")
            self._copy_worksheet_with_formatting(template_instructions_ws, ws_instructions)
            
            # Save the file
            wb.save(output_file)
            logger.info(f"Created Instructions-only file: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating Instructions-only file {output_file}: {e}")
            return False
    
    def create_rbmf1_tab(self, source_file: Path) -> pd.DataFrame:
        """Create RBMF_1 tab with additional columns from source file.
        
        Args:
            source_file: Path to the source file
            
        Returns:
            DataFrame with RBMF_1 data including new columns
        """
        try:
            # Read the RBMF tab from source file
            rbmf_df = pd.read_excel(source_file, sheet_name='RBMF')
            # Track original row order to preserve through transformations
            rbmf_df['__source_row_index__'] = rbmf_df.index
            
            if rbmf_df.empty:
                logger.warning(f"RBMF tab is empty in {source_file.name}")
                return pd.DataFrame()
            
            logger.info(f"Loaded RBMF data with {len(rbmf_df)} rows from {source_file.name}")
            
            # Check if required column exists
            if 'Reporting Year - Quarter' not in rbmf_df.columns:
                logger.error(f"Required column 'Reporting Year - Quarter' not found in {source_file.name}")
                return pd.DataFrame()
            
            # Create a copy to work with
            rbmf1_df = rbmf_df.copy()
            
            # Parse Year and Quarter from "Reporting Year - Quarter" column
            rbmf1_df['Year'] = None
            rbmf1_df['Quarter'] = None
            rbmf1_df['Half Year'] = None
            
            for idx, row in rbmf1_df.iterrows():
                try:
                    year_quarter_str = str(row['Reporting Year - Quarter']).strip()
                    
                    # Parse year and quarter (handle various formats)
                    if '-' in year_quarter_str:
                        parts = year_quarter_str.split('-')
                        if len(parts) >= 2:
                            year = parts[0].strip()
                            quarter = parts[1].strip()
                            
                            # Clean quarter (remove 'Q' prefix if present)
                            quarter = quarter.replace('Q', '').strip()
                            
                            try:
                                year_int = int(year)
                                quarter_int = int(quarter)
                                
                                rbmf1_df.at[idx, 'Year'] = year_int
                                rbmf1_df.at[idx, 'Quarter'] = quarter_int
                                
                                # Determine Half Year
                                if quarter_int < 3:
                                    rbmf1_df.at[idx, 'Half Year'] = 'H1'
                                else:
                                    rbmf1_df.at[idx, 'Half Year'] = 'H2'
                                    
                            except ValueError:
                                logger.warning(f"Invalid year/quarter format in row {idx}: {year_quarter_str}")
                                continue
                        else:
                            logger.warning(f"Invalid year-quarter format in row {idx}: {year_quarter_str}")
                            continue
                    else:
                        logger.warning(f"No dash found in year-quarter string in row {idx}: {year_quarter_str}")
                        continue
                        
                except Exception as e:
                    logger.warning(f"Error parsing year-quarter in row {idx}: {e}")
                    continue
            
            # Add ID column (Indicator_ID with spaces removed)
            if 'Indicator_ID' in rbmf1_df.columns:
                rbmf1_df['ID'] = rbmf1_df['Indicator_ID'].astype(str).str.replace(' ', '')
            else:
                logger.warning(f"Indicator_ID column not found in {source_file.name}")
                rbmf1_df['ID'] = ''
            
            # Reorder columns: put new columns after "Reporting Year - Quarter"
            columns = list(rbmf1_df.columns)
            
            # Remove any duplicate columns that might have been created
            rbmf1_df = rbmf1_df.loc[:, ~rbmf1_df.columns.duplicated()]
            columns = list(rbmf1_df.columns)
            
            # Find the position of "Reporting Year - Quarter"
            try:
                reporting_pos = columns.index('Reporting Year - Quarter')
                
                # Insert new columns after "Reporting Year - Quarter"
                new_columns = columns[:reporting_pos + 1] + ['Year', 'Quarter', 'Half Year'] + columns[reporting_pos + 1:]
                
                # Insert ID column after Indicator_ID if it exists
                if 'Indicator_ID' in new_columns:
                    indicator_id_pos = new_columns.index('Indicator_ID')
                    new_columns = new_columns[:indicator_id_pos + 1] + ['ID'] + new_columns[indicator_id_pos + 1:]
                else:
                    # If Indicator_ID doesn't exist, add ID at the end
                    new_columns.append('ID')
                
                # Remove duplicates from new_columns list
                seen = set()
                unique_columns = []
                for col in new_columns:
                    if col not in seen:
                        seen.add(col)
                        unique_columns.append(col)
                
                rbmf1_df = rbmf1_df[unique_columns]
                
            except ValueError:
                logger.warning(f"'Reporting Year - Quarter' column not found for reordering in {source_file.name}")
            
            # Apply column mapping if available
            column_mappings = self.load_column_mapping()
            if column_mappings:
                logger.info("Applying column mappings to RBMF_1 data")
                rbmf1_df = self._apply_column_mapping(rbmf1_df, column_mappings)
            
            logger.info(f"Created RBMF_1 data with {len(rbmf1_df)} rows and {len(rbmf1_df.columns)} columns")
            return rbmf1_df
            
        except Exception as e:
            logger.error(f"Error creating RBMF_1 tab from {source_file.name}: {e}")
            return pd.DataFrame()
    
    def create_rbmf2_tab(self, rbmf1_df: pd.DataFrame) -> pd.DataFrame:
        """Create RBMF_2 tab with aggregated data from RBMF_1.
        
        Args:
            rbmf1_df: DataFrame from RBMF_1 tab
            
        Returns:
            DataFrame with aggregated RBMF_2 data
        """
        try:
            if rbmf1_df.empty:
                logger.warning("RBMF_1 data is empty, cannot create RBMF_2")
                return pd.DataFrame()
            
            # Create a copy to work with
            rbmf2_df = rbmf1_df.copy()
            
            # Transform Half Year to include year prefix (e.g., "2024-H1")
            rbmf2_df['Half Year'] = rbmf2_df['Year'].astype(str) + '-' + rbmf2_df['Half Year']
            
            # Handle empty/null values in Completed Output Number as 0
            rbmf2_df['Completed Output Number'] = pd.to_numeric(
                rbmf2_df['Completed Output Number'], errors='coerce'
            ).fillna(0)
            
            # Define grouping columns
            group_columns = ['ID', 'Half Year', 'Year']
            
            # Define columns that should be merged (may have different values)
            merge_columns = ['Reporting Year - Quarter', 'Quarter']
            
            # Define columns that should be the same (take first value), excluding custom-rule columns
            custom_rule_columns = ['Project Output', 'Project Output Status']
            other_columns = [col for col in rbmf2_df.columns 
                           if col not in group_columns + merge_columns + ['Completed Output Number'] + custom_rule_columns]
            
            # Perform aggregation
            agg_dict = {}
            
            # For merge columns, concatenate unique values with semicolon
            for col in merge_columns:
                if col in rbmf2_df.columns:
                    agg_dict[col] = lambda x: '; '.join(x.dropna().astype(str).unique())
            
            # For Completed Output Number, sum and rename
            if 'Completed Output Number' in rbmf2_df.columns:
                agg_dict['Completed Output Number'] = 'sum'
            
            # For other columns, take first value (they should be the same)
            for col in other_columns:
                if col in rbmf2_df.columns:
                    agg_dict[col] = 'first'
            
            # Preserve __source_row_index__ by taking the minimum (earliest appearance)
            if '__source_row_index__' in rbmf2_df.columns:
                agg_dict['__source_row_index__'] = 'min'

            # Custom aggregation for Project Output: merge distinct values with dash bullets on new lines
            if 'Project Output' in rbmf2_df.columns:
                def _merge_project_outputs(series: pd.Series) -> str:
                    values = [str(v).strip() for v in series.dropna().astype(str) if str(v).strip() != '']
                    unique_values = list(dict.fromkeys(values))  # preserve order, remove duplicates
                    if len(unique_values) <= 1:
                        return unique_values[0] if unique_values else ''
                    # Join as dash bullets separated by new lines
                    return "\n".join(f"- {v}" for v in unique_values)
                agg_dict['Project Output'] = _merge_project_outputs

            # Custom aggregation for Project Output Status with priority
            if 'Project Output Status' in rbmf2_df.columns:
                def _aggregate_status(series: pd.Series) -> str:
                    vals = [str(v).strip() for v in series.dropna().astype(str)]
                    # Priority: Completed > In Progress > others (first occurrence)
                    if any(v.lower() == 'completed' for v in vals):
                        return 'Completed'
                    if any(v.lower() == 'in progress' for v in vals):
                        return 'In Progress'
                    return vals[0] if len(vals) > 0 else ''
                agg_dict['Project Output Status'] = _aggregate_status

            # Custom aggregation for Progress Notes/Comments: merge distinct values with dash bullets
            if 'Progress Notes/Comments' in rbmf2_df.columns:
                def _merge_progress_notes(series: pd.Series) -> str:
                    values = [str(v).strip() for v in series.dropna().astype(str) if str(v).strip() != '']
                    unique_values = list(dict.fromkeys(values))  # preserve order, remove duplicates
                    if len(unique_values) <= 1:
                        return unique_values[0] if unique_values else ''
                    # Join as dash bullets separated by new lines
                    return "\n".join(f"- {v}" for v in unique_values)
                agg_dict['Progress Notes/Comments'] = _merge_progress_notes

            # Custom aggregation for Supporting Document: merge distinct values with dash bullets
            if 'Supporting Document' in rbmf2_df.columns:
                def _merge_supporting_docs(series: pd.Series) -> str:
                    values = [str(v).strip() for v in series.dropna().astype(str) if str(v).strip() != '']
                    unique_values = list(dict.fromkeys(values))  # preserve order, remove duplicates
                    if len(unique_values) <= 1:
                        return unique_values[0] if unique_values else ''
                    # Join as dash bullets separated by new lines
                    return "\n".join(f"- {v}" for v in unique_values)
                agg_dict['Supporting Document'] = _merge_supporting_docs
            
            # Group and aggregate
            # Preserve original appearance order of groups
            rbmf2_aggregated = rbmf2_df.groupby(group_columns, dropna=False, sort=False).agg(agg_dict).reset_index()
            
            # Rename Completed Output Number to Total Completed Output Number
            if 'Completed Output Number' in rbmf2_aggregated.columns:
                rbmf2_aggregated = rbmf2_aggregated.rename(
                    columns={'Completed Output Number': 'Total Completed Output Number'}
                )
            
            # Validate data consistency for non-grouping columns (exclude fields with custom aggregation)
            custom_agg_fields = ['Project Output', 'Project Output Status', 'Progress Notes/Comments', 'Supporting Document']
            validation_columns = [col for col in other_columns if col not in custom_agg_fields]
            if validation_columns:  # Only validate if there are columns to validate
                self._validate_group_consistency(rbmf2_df, group_columns, validation_columns)
            
            # Sort aggregated rows by original appearance if available
            if '__source_row_index__' in rbmf2_aggregated.columns:
                rbmf2_aggregated = rbmf2_aggregated.sort_values(by='__source_row_index__', kind='stable').reset_index(drop=True)
            logger.info(f"Created RBMF_2 data with {len(rbmf2_aggregated)} rows and {len(rbmf2_aggregated.columns)} columns")
            return rbmf2_aggregated
            
        except Exception as e:
            logger.error(f"Error creating RBMF_2 tab: {e}")
            return pd.DataFrame()
    
    def _validate_group_consistency(self, df: pd.DataFrame, group_columns: list, other_columns: list) -> None:
        """Validate that non-grouping columns have consistent values within groups.
        
        Args:
            df: DataFrame to validate
            group_columns: Columns used for grouping
            other_columns: Columns that should have consistent values within groups
        """
        try:
            for col in other_columns:
                if col not in df.columns:
                    continue
                
                # Group by the grouping columns and check for inconsistencies
                grouped = df.groupby(group_columns)[col].apply(lambda x: x.nunique())
                inconsistent_groups = grouped[grouped > 1]
                
                if len(inconsistent_groups) > 0:
                    logger.warning(f"Data inconsistency found in column '{col}': {len(inconsistent_groups)} groups have different values")
                    for group_key, count in inconsistent_groups.items():
                        logger.warning(f"  Group {group_key}: {count} different values")
                        
        except Exception as e:
            logger.warning(f"Error validating group consistency: {e}")
    
    def create_rbmf_final_tab(self, rbmf2_df: pd.DataFrame, source_file_name: str = None) -> pd.DataFrame:
        """Create RBMF_Final tab with template structure and RBMF_2 data mapping.
        
        Args:
            rbmf2_df: DataFrame from RBMF_2 tab
            source_file_name: Name of the source file for Project ID mapping
            
        Returns:
            DataFrame with RBMF_Final data mapped from RBMF_2
        """
        try:
            if rbmf2_df.empty:
                logger.warning("RBMF_2 data is empty, cannot create RBMF_Final")
                return pd.DataFrame()
            
            # Load template RBMF structure
            template_rbmf_df = self._load_template_rbmf_structure()
            if template_rbmf_df.empty:
                logger.error("Could not load template RBMF structure")
                return pd.DataFrame()
            
            # Define mapping rules (RBMF_Final column <- RBMF_2 column)
            mapping_rules = {
                'Strategic Outcome': 'Primary Outcome Area',
                'Indicator category': 'Result_Type_Data',
                'Indicator name': 'Indicators',
                'Indicator Description': 'Project Output',
                'Periodical Target': 'Output Target Number',
                'Target Reporting Cycle': 'Half Year',
                'Indicator Status': 'Project Output Status',
                'Periodical Result': 'Total Completed Output Number',
                'Result Notes/Comments': 'Progress Notes/Comments',
                'Supporting Document': 'Supporting Document',
                'Country': 'Country',
                'Indicator ID': 'ID',  # Map from ID field in RBMF_2
                'Project ID': None  # Will be set based on file mapping
            }
            
            # Create RBMF_Final DataFrame with template structure
            rbmf_final_df = pd.DataFrame()
            
            # Map data from RBMF_2 to RBMF_Final
            for final_col, source_col in mapping_rules.items():
                if final_col in template_rbmf_df.columns:
                    if source_col is None:
                        # Set Project ID based on file mapping, other None fields to empty
                        if final_col == 'Project ID':
                            if source_file_name:
                                project_id = self.get_project_id_for_file(source_file_name)
                                rbmf_final_df[final_col] = project_id
                                logger.info(f"Set {final_col} to '{project_id}' from file mapping")
                            else:
                                rbmf_final_df[final_col] = "NOT FOUND"
                                logger.warning(f"Set {final_col} to 'NOT FOUND' - no source file name provided")
                        else:
                            rbmf_final_df[final_col] = ''
                            logger.info(f"Set {final_col} to empty as specified")
                    elif source_col in rbmf2_df.columns:
                        if final_col == 'Target Reporting Cycle':
                            # Remove "-" character from Half Year
                            rbmf_final_df[final_col] = rbmf2_df[source_col].astype(str).str.replace('-', '')
                        else:
                            rbmf_final_df[final_col] = rbmf2_df[source_col]
                        logger.info(f"Mapped {final_col} <- {source_col}")
                    else:
                        logger.warning(f"Source column '{source_col}' not found in RBMF_2 data")
                        rbmf_final_df[final_col] = ''
                else:
                    logger.warning(f"Template column '{final_col}' not found in template RBMF structure")
            
            # Ensure same number of rows as RBMF_2
            if len(rbmf_final_df) != len(rbmf2_df):
                logger.warning(f"Row count mismatch: RBMF_Final has {len(rbmf_final_df)} rows, RBMF_2 has {len(rbmf2_df)} rows")
            
            # Apply column mapping if available
            column_mappings = self.load_column_mapping()
            if column_mappings:
                logger.info("Applying column mappings to RBMF_Final data")
                rbmf_final_df = self._apply_column_mapping(rbmf_final_df, column_mappings)
            
            # Adjust Indicator category based on Indicator ID prefix
            try:
                if 'Indicator ID' in rbmf_final_df.columns and 'Indicator category' in rbmf_final_df.columns:
                    indicator_id_series = rbmf_final_df['Indicator ID'].astype(str).fillna('')
                    # Create masks
                    op_mask = indicator_id_series.str.startswith('OP', na=False)
                    oc_mask = indicator_id_series.str.startswith('OC', na=False)
                    in_mask = indicator_id_series.str.startswith('In', na=False)
                    # Apply updates without disturbing other values
                    if op_mask.any():
                        rbmf_final_df.loc[op_mask, 'Indicator category'] = 'Output'
                    if oc_mask.any():
                        rbmf_final_df.loc[oc_mask, 'Indicator category'] = 'Outcome'
                    if in_mask.any():
                        rbmf_final_df.loc[in_mask, 'Indicator category'] = 'Outcome'
                    logger.info("Updated Indicator category from Indicator ID prefixes (OP->Output, OC/In->Outcome)")
                else:
                    logger.warning("Cannot adjust Indicator category: required columns missing")
            except Exception as e:
                logger.warning(f"Failed to adjust Indicator category from Indicator ID: {e}")

            # Preserve original row order in RBMF_Final when available
            if '__source_row_index__' in rbmf2_df.columns:
                rbmf_final_df['__source_row_index__'] = rbmf2_df['__source_row_index__']
                rbmf_final_df = rbmf_final_df.sort_values(by='__source_row_index__', kind='stable').reset_index(drop=True)
            logger.info(f"Created RBMF_Final data with {len(rbmf_final_df)} rows and {len(rbmf_final_df.columns)} columns")
            return rbmf_final_df
            
        except Exception as e:
            logger.error(f"Error creating RBMF_Final tab: {e}")
            return pd.DataFrame()
    
    def _load_template_rbmf_structure(self) -> pd.DataFrame:
        """Load the RBMF tab structure from the template file.
        
        Returns:
            DataFrame with template RBMF column structure
        """
        try:
            # Read the RBMF tab from template to get column structure
            template_rbmf_df = pd.read_excel(self.template_file, sheet_name='RBMF')
            logger.info(f"Loaded template RBMF structure with {len(template_rbmf_df.columns)} columns")
            return template_rbmf_df
            
        except Exception as e:
            logger.error(f"Error loading template RBMF structure: {e}")
            return pd.DataFrame()
    
    def create_output_file(self, source_file: Path, output_file: Path, apply_filter: bool = False) -> bool:
        """Create output file with Instructions and RBMF_Final tabs, optionally including intermediate steps.
        
        Args:
            source_file: Path to the source file
            output_file: Path for the output file
            apply_filter: Whether to apply filtering to RBMF tab data
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if self.template_instructions is None:
                raise ValueError("Template instructions not loaded. Call load_template_instructions() first.")
            
            # Load file-to-project mapping if not already loaded
            if self.file_to_project_id_mapping is None:
                self.load_file_to_project_id_mapping()
            
            # Ensure output file has .xlsx extension for Google Sheets compatibility
            if not output_file.suffix.lower() == '.xlsx':
                output_file = output_file.with_suffix('.xlsx')
                logger.info(f"Changed output file extension to .xlsx for Google Sheets compatibility: {output_file.name}")
            
            # Load template workbook to preserve formatting
            template_wb = self.load_template_workbook()
            
            # Always create RBMF_1 data (needed for RBMF_Final)
            rbmf1_df = self.create_rbmf1_tab(source_file)
            if rbmf1_df.empty:
                logger.warning(f"No RBMF_1 data created for {source_file.name}")
                return False
            
            # Create RBMF_2 data from RBMF_1 (needed for RBMF_Final)
            rbmf2_df = self.create_rbmf2_tab(rbmf1_df)
            if rbmf2_df.empty:
                logger.warning(f"No RBMF_2 data created for {source_file.name}")
                return False
            
            # Create RBMF_Final data from RBMF_2
            rbmf_final_df = self.create_rbmf_final_tab(rbmf2_df, source_file.name)
            if rbmf_final_df.empty:
                logger.warning(f"No RBMF_Final data created for {source_file.name}")
                return False
            
            # Apply filtering if requested
            if apply_filter:
                rbmf_final_df = self._apply_rbmf_filtering(rbmf_final_df)
                logger.info(f"Applied filtering to RBMF_Final data: {len(rbmf_final_df)} rows remaining")
            
            # Extract header style from source RBMF sheet (if available)
            header_style = self._extract_rbmf_header_style(source_file)
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Copy Instructions tab with all formatting from template
            template_instructions_ws = template_wb['Instructions']
            ws_instructions = wb.create_sheet("Instructions")
            self._copy_worksheet_with_formatting(template_instructions_ws, ws_instructions)
            
            # Add intermediate tabs only if include_steps is True
            if self.include_steps:
                # Add RBMF_1 tab
                ws_rbmf1 = wb.create_sheet("RBMF_1")
                self._write_dataframe_to_worksheet(rbmf1_df, ws_rbmf1, header_style=header_style)
                
                # Add RBMF_2 tab
                ws_rbmf2 = wb.create_sheet("RBMF_2")
                self._write_dataframe_to_worksheet(rbmf2_df, ws_rbmf2, header_style=header_style)
            
            # Always add RBMF tab with template formatting
            ws_rbmf = wb.create_sheet("RBMF")
            self._write_dataframe_to_worksheet_with_template_formatting(
                rbmf_final_df, ws_rbmf, template_wb
            )
            
            # Copy Overview tab from SOURCE (preferred); fallback to template if missing
            copied_overview = False
            
            # Copy all other tabs from source file except RBMF and Instructions
            try:
                # Handle files without supported extension by copying to a temp .xlsx
                path_to_open = Path(source_file)
                temp_created_for_copy = False
                if path_to_open.suffix.lower() not in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
                    import shutil, tempfile
                    tmp_dir = tempfile.gettempdir()
                    temp_path = Path(tmp_dir) / "_rbmf_copy_tabs_probe.xlsx"
                    shutil.copy(path_to_open, temp_path)
                    path_to_open = temp_path
                    temp_created_for_copy = True

                source_wb = openpyxl.load_workbook(path_to_open)
                # Transform Overview tab using 5-table mapping algorithm
                try:
                    copied_overview = self._transform_overview_tab(source_wb, template_wb, wb, rbmf_final_df)
                    if copied_overview:
                        logger.info("Successfully transformed Overview tab using 5-table mapping algorithm")
                    else:
                        logger.warning("Failed to transform Overview tab")
                except Exception as e:
                    logger.warning(f"Failed to transform Overview tab: {e}")
                    copied_overview = False

                for sheet_name in source_wb.sheetnames:
                    # Skip RBMF and Instructions. Always skip Overview (we copy it from template)
                    if sheet_name in ['RBMF', 'Instructions', 'Overview']:
                        continue
                    if sheet_name in wb.sheetnames:
                        logger.warning(f"Skipping source tab due to name conflict: {sheet_name}")
                        continue
                    try:
                        source_ws = source_wb[sheet_name]
                        target_ws = wb.create_sheet(sheet_name)
                        self._copy_worksheet_with_formatting(source_ws, target_ws)
                        logger.info(f"Copied source tab '{sheet_name}' with formatting")
                    except Exception as e:
                        logger.warning(f"Failed to copy source tab '{sheet_name}': {e}")
                try:
                    source_wb.close()
                except Exception:
                    pass
                finally:
                    if temp_created_for_copy and path_to_open.exists():
                        try:
                            path_to_open.unlink()
                        except Exception:
                            pass
            except Exception as e:
                logger.warning(f"Could not open source workbook to copy tabs: {e}")
            
            # Optimize formatting for Google Sheets compatibility (skip Overview to preserve exact formatting)
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Overview':
                    continue
                self._optimize_formatting_for_google_sheets(wb[sheet_name])
            
            # Save the file
            wb.save(output_file)
            
            # Validate Google Sheets compatibility
            self._validate_google_sheets_compatibility(output_file)
            
            if self.include_steps:
                logger.info(f"Created file with Instructions, RBMF_1, RBMF_2, and RBMF tabs: {output_file}")
            else:
                logger.info(f"Created file with Instructions and RBMF tabs: {output_file}")
            return True
            
        except Exception as e:
            mode_desc = "with intermediate steps" if self.include_steps else "final only"
            logger.error(f"Error creating {mode_desc} file {output_file}: {e}")
            return False

    def _extract_rbmf_header_style(self, source_file: Path) -> Dict[str, Any]:
        """Extract a representative header cell style from the source RBMF sheet.
        
        Returns a dict with keys: font, fill, border, alignment, number_format, or empty dict if unavailable.
        """
        try:
            path_to_open = Path(source_file)
            temp_created = False
            # If no supported extension, copy to a temp .xlsx so openpyxl can open it
            if path_to_open.suffix.lower() not in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
                import shutil, tempfile
                tmp_dir = tempfile.gettempdir()
                temp_path = Path(tmp_dir) / "_rbmf_header_probe.xlsx"
                shutil.copy(path_to_open, temp_path)
                path_to_open = temp_path
                temp_created = True
            
            wb = openpyxl.load_workbook(path_to_open)
            if 'RBMF' not in wb.sheetnames:
                wb.close()
                if temp_created and path_to_open.exists():
                    try:
                        path_to_open.unlink()
                    except Exception:
                        pass
                return {}
            ws = wb['RBMF']
            # Use first row, first non-empty cell style as template
            for cell in ws[1]:
                if cell.value is not None:
                    style = {
                        'font': copy(cell.font) if cell.has_style else None,
                        'fill': copy(cell.fill) if cell.has_style else None,
                        'border': copy(cell.border) if cell.has_style else None,
                        'alignment': copy(cell.alignment) if cell.has_style else None,
                        'number_format': copy(cell.number_format) if cell.has_style else None,
                    }
                    wb.close()
                    if temp_created and path_to_open.exists():
                        try:
                            path_to_open.unlink()
                        except Exception:
                            pass
                    return style
            wb.close()
            if temp_created and path_to_open.exists():
                try:
                    path_to_open.unlink()
                except Exception:
                    pass
            return {}
        except Exception as e:
            logger.warning(f"Could not extract RBMF header style from {source_file.name}: {e}")
            return {}

    def _copy_worksheet_with_formatting(self, source_ws, target_ws) -> None:
        """Copy worksheet with all formatting preserved.
        
        Args:
            source_ws: Source worksheet (from template)
            target_ws: Target worksheet (new file)
        """
        # Copy all cells with their formatting
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                
                # Copy cell formatting
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # Copy column dimensions
        for column_letter in source_ws.column_dimensions:
            target_ws.column_dimensions[column_letter] = copy(source_ws.column_dimensions[column_letter])
        
        # Copy row dimensions
        for row_number in source_ws.row_dimensions:
            target_ws.row_dimensions[row_number] = copy(source_ws.row_dimensions[row_number])
        
        # Copy merged cells
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))
        
        # Copy worksheet properties (only if they exist and can be copied)
        if hasattr(source_ws, 'sheet_properties'):
            target_ws.sheet_properties = copy(source_ws.sheet_properties)
        
        # Note: sheet_view and protection properties are read-only in openpyxl
        # so we skip copying them to avoid errors
        
        try:
            title = getattr(source_ws, 'title', 'worksheet')
        except Exception:
            title = 'worksheet'
        logger.info(f"Copied '{title}' tab with formatting")
    
    def _clear_cell_range(self, worksheet, cell_range: str) -> None:
        """Clear data and colors from a specific cell range.
        
        Args:
            worksheet: The worksheet to modify
            cell_range: Cell range in format like 'E8:H11'
        """
        try:
            from openpyxl.utils import range_boundaries
            
            # Parse the cell range
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            
            # First, unmerge any merged cells in the range
            merged_ranges_to_remove = []
            for merged_range in worksheet.merged_cells.ranges:
                # Check if the merged range overlaps with our target range
                if (merged_range.min_row <= max_row and merged_range.max_row >= min_row and
                    merged_range.min_col <= max_col and merged_range.max_col >= min_col):
                    merged_ranges_to_remove.append(merged_range)
            
            # Remove overlapping merged cells
            for merged_range in merged_ranges_to_remove:
                worksheet.unmerge_cells(str(merged_range))
            
            # Clear cells in the range
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    # Skip if it's a merged cell (shouldn't happen after unmerging, but safety check)
                    if hasattr(cell, 'value') and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.value = None
                        # Reset to default formatting (no fill, no font color)
                        cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                        cell.font = openpyxl.styles.Font()
                        cell.border = openpyxl.styles.Border()
                        cell.alignment = openpyxl.styles.Alignment()
                        cell.number_format = 'General'
            
            logger.info(f"Cleared data and colors from cell range {cell_range}")
            
        except Exception as e:
            logger.error(f"Error clearing cell range {cell_range}: {e}")
            raise
    
    def _hide_borders_on_empty_cells(self, worksheet) -> None:
        """Hide borders on cells that don't contain data and set borders for stakeholder table.
        
        Args:
            worksheet: The worksheet to modify
        """
        try:
            from openpyxl.styles import Border, Side
            
            # Create a very light border (almost invisible)
            light_border = Border(
                left=Side(style='thin', color='FFFFFF'),  # White color
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )
            
            # Create a standard border for the stakeholder table
            table_border = Border(
                left=Side(style='thin', color='000000'),  # Black color, thin thickness
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Track how many cells were modified
            modified_cells = 0
            table_cells = 0
            
            # Get merged cell ranges to avoid them
            merged_ranges = set()
            for merged_range in worksheet.merged_cells.ranges:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_ranges.add((row, col))
            
            # Find the stakeholder table
            stakeholder_table_range = self._find_stakeholder_table(worksheet)
            
            # Iterate through all cells in the worksheet
            for row in worksheet.iter_rows():
                for cell in row:
                    # Skip merged cells
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    
                    # Skip if this cell is part of a merged range
                    if (cell.row, cell.column) in merged_ranges:
                        continue
                    
                    # Check if this cell is in the stakeholder table
                    is_in_table = False
                    if stakeholder_table_range:
                        min_row, min_col, max_row, max_col = stakeholder_table_range
                        if (min_row <= cell.row <= max_row and min_col <= cell.column <= max_col):
                            is_in_table = True
                            # Only set borders for cells with data in the table
                            has_data = (cell.value is not None and 
                                       not (isinstance(cell.value, str) and cell.value.strip() == ''))
                            if has_data:
                                cell.border = table_border
                                table_cells += 1
                    
                    # For cells not in the table, apply the original logic
                    if not is_in_table:
                        # Check if cell is empty (no value or empty string)
                        is_empty = (cell.value is None or 
                                   (isinstance(cell.value, str) and cell.value.strip() == ''))
                        
                        if is_empty:
                            # Set very light borders on empty cells
                            cell.border = light_border
                            modified_cells += 1
            
            logger.info(f"Hidden borders on {modified_cells} empty cells and set borders on {table_cells} data cells in stakeholder table")
            
        except Exception as e:
            logger.error(f"Error hiding borders on empty cells: {e}")
            raise
    
    def _transform_overview_tab(self, source_wb, template_wb, wb, rbmf_final_df=None):
        """
        Transform Overview tab from source to 2025 template format using 5-table mapping algorithm.
        
        Args:
            source_wb: Source workbook (original RBMF file)
            template_wb: Template workbook (2025 template)
            wb: Target workbook (final output file)
            rbmf_final_df: DataFrame with RBMF_Final data for Project ID and Strategic Outcome mapping
        
        Returns:
            bool: True if Overview tab was successfully created, False otherwise
        """
        try:
            # Check if source has Overview tab
            if 'Overview' not in source_wb.sheetnames:
                logger.warning("Source file does not contain Overview tab")
                return False
            
            # Check if template has Overview tab
            if 'Overview' not in template_wb.sheetnames:
                logger.warning("Template file does not contain Overview tab")
                return False
            
            # Read Overview data from source and template worksheets
            source_ws = source_wb['Overview']
            template_ws = template_wb['Overview']
            
            # Convert worksheets to DataFrames
            source_data = []
            for row in source_ws.iter_rows(values_only=True):
                source_data.append(row)
            source_df = pd.DataFrame(source_data)
            
            template_data = []
            for row in template_ws.iter_rows(values_only=True):
                template_data.append(row)
            template_df = pd.DataFrame(template_data)
            
            logger.info("Starting Overview tab transformation using 5-table mapping algorithm")
            
            # Create output dataframe starting with template
            df_output = template_df.copy()
            
            # Note: Dynamic row expansion will be handled in _map_table_5_stakeholder_quotes
            # based on actual stakeholder count found in the original file
            
            # Map all 5 tables
            self._map_table_1_general_overview(source_df, df_output)
            self._map_table_2_project_overview(source_df, df_output)
            self._map_table_3_implementation_partner(source_df, df_output)
            self._map_table_4_project_stakeholders(source_df, df_output)
            self._map_table_5_stakeholder_quotes(source_df, df_output)
            
            # Map RBMF data to Overview tab fields
            if rbmf_final_df is not None and not rbmf_final_df.empty:
                self._map_rbmf_data_to_overview(rbmf_final_df, df_output)
            
            # Create Overview sheet in target workbook
            ws_overview = wb.create_sheet("Overview")
            
            # Write transformed data to the sheet
            for row_idx, row in df_output.iterrows():
                for col_idx, value in enumerate(row):
                    if pd.notna(value) and str(value).strip():
                        ws_overview.cell(row=row_idx + 1, column=col_idx + 1, value=str(value))
            
            # Apply formatting from template (preserve exact template layout)
            self._copy_worksheet_formatting(template_wb['Overview'], ws_overview)

            # Ensure all cells wrap text as requested
            self._set_wrap_text_for_all_cells(ws_overview)

            # After copying template formatting, hide borders on empty cells to match template's emerging tables look
            self._hide_borders_on_empty_cells(ws_overview)

            # Note: We no longer copy the template Overview as a separate 'Overview 2025' sheet
            
            logger.info("Successfully transformed Overview tab using 5-table mapping algorithm")
            return True
            
        except Exception as e:
            logger.error(f"Failed to transform Overview tab: {e}")
            return False
    
    def _map_table_1_general_overview(self, df_original: pd.DataFrame, df_output: pd.DataFrame):
        """Map Table 1: General Overview Section (Left Side)"""
        try:
            # Row 0: Section header
            df_output.iloc[0, 0] = df_original.iloc[0, 0]  # "General Overview"
            
            # Row 1: Country
            df_output.iloc[1, 0] = "Country"
            df_output.iloc[1, 1] = df_original.iloc[2, 1]  # "Indonesia"
            
            # Row 2: Project Name
            df_output.iloc[2, 0] = "Project Name"
            df_output.iloc[2, 1] = df_original.iloc[3, 1]  # Full project name
            
            # Row 3: Project ID
            df_output.iloc[3, 0] = "Project ID"
            df_output.iloc[3, 1] = ""  # Not present in original
            
            # Row 4: Implementing Partner
            df_output.iloc[4, 0] = "Implementing Partner/Retainer Name"
            df_output.iloc[4, 1] = df_original.iloc[1, 1]  # "Aquatera"
            
            # Row 5: Implementation period
            df_output.iloc[5, 0] = "Implementation period"
            period = df_original.iloc[5, 1]  # "July 2024 - November 2025"
            if ' - ' in period:
                from_date, to_date = period.split(' - ')
                df_output.iloc[5, 1] = from_date.strip()
                df_output.iloc[5, 2] = to_date.strip()
            
            # Row 6: Post Implementation Monitoring
            df_output.iloc[6, 0] = "Post Implementation Monitoring"
            df_output.iloc[6, 1] = ""  # Not present in original
            
            logger.info("✓ Table 1 (General Overview) mapped successfully")
        except Exception as e:
            logger.warning(f"Failed to map Table 1: {e}")
    
    def _map_table_2_project_overview(self, df_original: pd.DataFrame, df_output: pd.DataFrame):
        """Map Table 2: Project Overview Section (Right Side)"""
        try:
            # Row 0: Section header
            df_output.iloc[0, 4] = df_original.iloc[0, 4]  # "Project overview"
            
            # Row 1: Primary strategic outcome
            df_output.iloc[1, 4] = "Primary strategic outcome"
            df_output.iloc[1, 5] = df_original.iloc[1, 5]  # "SO4 - Knowledge and Awareness Building"
            
            # Row 2: Impact
            df_output.iloc[2, 4] = "Impact"
            df_output.iloc[2, 5] = df_original.iloc[2, 5]  # Impact description
            
            # Row 3: Outcome
            df_output.iloc[3, 4] = "Outcome"
            df_output.iloc[3, 5] = df_original.iloc[3, 5]  # Outcome description
            
            # Row 4: Output
            df_output.iloc[4, 4] = "Output"
            df_output.iloc[4, 5] = df_original.iloc[4, 5]  # Output description
            
            logger.info("✓ Table 2 (Project Overview) mapped successfully")
        except Exception as e:
            logger.warning(f"Failed to map Table 2: {e}")
    
    def _map_table_3_implementation_partner(self, df_original: pd.DataFrame, df_output: pd.DataFrame):
        """Map Table 3: Implementation Partner Overview Table"""
        try:
            # Row 8: Section header
            df_output.iloc[8, 0] = df_original.iloc[7, 0]  # "Implementation Partner Overview"
            
            # Row 9: Table header
            df_output.iloc[9, 0] = df_original.iloc[8, 0]  # "Implementation Partner"
            df_output.iloc[9, 1] = df_original.iloc[8, 1]  # "Total"
            df_output.iloc[9, 2] = df_original.iloc[8, 2]  # "Female"
            
            # Row 10: Project Team
            df_output.iloc[10, 0] = df_original.iloc[9, 0]  # "Project Team"
            df_output.iloc[10, 1] = df_original.iloc[9, 1]  # "17"
            df_output.iloc[10, 2] = df_original.iloc[9, 2]  # "9"
            
            # Row 11: Number of Founders
            df_output.iloc[11, 0] = df_original.iloc[10, 0]  # "Number of Founders"
            df_output.iloc[11, 1] = df_original.iloc[10, 1]  # "1"
            df_output.iloc[11, 2] = df_original.iloc[10, 2]  # "0"
            
            # Row 12: Senior Management
            df_output.iloc[12, 0] = df_original.iloc[11, 0]  # "Senior Management"
            df_output.iloc[12, 1] = df_original.iloc[11, 1]  # "5 Directors"
            df_output.iloc[12, 2] = df_original.iloc[11, 2]  # "2"
            
            logger.info("✓ Table 3 (Implementation Partner) mapped successfully")
        except Exception as e:
            logger.warning(f"Failed to map Table 3: {e}")
    
    def _map_table_4_project_stakeholders(self, df_original: pd.DataFrame, df_output: pd.DataFrame):
        """Map Table 4: Project Stakeholders Table"""
        try:
            # Row 14: Section header
            df_output.iloc[14, 0] = df_original.iloc[13, 0]  # "Project Stakeholders"
            df_output.iloc[14, 4] = df_original.iloc[13, 4]  # "Project Stakeholders/Beneficiary Quotes"
            
            # Row 15: Table header
            df_output.iloc[15, 0] = df_original.iloc[14, 0]  # "Stakeholder Name"
            df_output.iloc[15, 1] = df_original.iloc[14, 1]  # "Department/Organisation Name"
            df_output.iloc[15, 2] = df_original.iloc[14, 2]  # "Position"
            df_output.iloc[15, 4] = df_original.iloc[14, 4]  # "Stakeholder/Beneficiary Name"
            df_output.iloc[15, 5] = df_original.iloc[14, 5]  # "Department/Organisation Name"
            df_output.iloc[15, 6] = df_original.iloc[14, 6]  # "Position"
            df_output.iloc[15, 7] = df_original.iloc[14, 7]  # "Quote"
            
            logger.info("✓ Table 4 (Project Stakeholders) mapped successfully")
        except Exception as e:
            logger.warning(f"Failed to map Table 4: {e}")
    
    def _map_table_5_stakeholder_quotes(self, df_original: pd.DataFrame, df_output: pd.DataFrame):
        """Map Table 5: Project Stakeholders/Beneficiary Quotes Table - DYNAMIC VERSION"""
        try:
            # Dynamically find stakeholder data range in original file
            stakeholder_rows = self._find_stakeholder_data_rows(df_original)
            
            if not stakeholder_rows:
                logger.warning("No stakeholder data found in original file")
                return
            
            logger.info(f"Found {len(stakeholder_rows)} stakeholder rows: {stakeholder_rows}")
            
            # Ensure we have enough rows in output template
            self._ensure_enough_template_rows(df_output, len(stakeholder_rows))
            
            # Copy stakeholder data rows to BOTH sides (left and right)
            template_row = 16  # Start after headers
            for orig_row in stakeholder_rows:
                if orig_row < len(df_original) and template_row < len(df_output):
                    # Copy stakeholder data to LEFT side (columns 0-2)
                    df_output.iloc[template_row, 0] = df_original.iloc[orig_row, 0]  # Name
                    df_output.iloc[template_row, 1] = df_original.iloc[orig_row, 1]  # Department
                    df_output.iloc[template_row, 2] = df_original.iloc[orig_row, 2]  # Position
                    
                    # Copy stakeholder data to RIGHT side (columns 4-6)
                    df_output.iloc[template_row, 4] = df_original.iloc[orig_row, 0]  # Name
                    df_output.iloc[template_row, 5] = df_original.iloc[orig_row, 1]  # Department
                    df_output.iloc[template_row, 6] = df_original.iloc[orig_row, 2]  # Position
                    df_output.iloc[template_row, 7] = df_original.iloc[orig_row, 7] if orig_row < len(df_original) and len(df_original.columns) > 7 else ""  # Quote
                    template_row += 1
            
            logger.info(f"✓ Table 5 (Stakeholder Quotes) mapped successfully - {len(stakeholder_rows)} stakeholders")
        except Exception as e:
            logger.warning(f"Failed to map Table 5: {e}")
    
    def _find_stakeholder_data_rows(self, df_original: pd.DataFrame) -> list:
        """Dynamically find rows containing stakeholder data in the original file."""
        stakeholder_rows = []
        
        try:
            # Look for stakeholder data starting from row 15 (after headers)
            for row_idx in range(15, len(df_original)):
                row = df_original.iloc[row_idx]
                
                # Check if this row has stakeholder data (name in column 0)
                name = row.iloc[0] if len(row) > 0 else None
                if pd.notna(name) and isinstance(name, str) and name.strip():
                    # Additional validation: check if it looks like a person's name
                    # (not a header, not empty, not a section title)
                    if not any(keyword in name.lower() for keyword in ['stakeholder', 'department', 'position', 'quote', 'project']):
                        stakeholder_rows.append(row_idx)
                    else:
                        # Stop if we hit another header/section
                        break
                else:
                    # Stop if we hit an empty row
                    break
            
            logger.info(f"Dynamic stakeholder detection found {len(stakeholder_rows)} stakeholder rows")
            return stakeholder_rows
            
        except Exception as e:
            logger.warning(f"Error in dynamic stakeholder detection: {e}")
            return []
    
    def _ensure_enough_template_rows(self, df_output: pd.DataFrame, stakeholder_count: int):
        """Ensure the template has enough rows for all stakeholders."""
        try:
            # Calculate how many rows we need (header + stakeholder rows)
            needed_rows = 16 + stakeholder_count  # 16 = rows 0-15 (headers and sections)
            
            # Add rows if needed
            while len(df_output) < needed_rows:
                import numpy as np
                new_row = pd.Series([np.nan] * len(df_output.columns))
                df_output = pd.concat([df_output, new_row.to_frame().T], ignore_index=True)
            
            logger.info(f"Ensured template has {len(df_output)} rows for {stakeholder_count} stakeholders")
            
        except Exception as e:
            logger.warning(f"Error ensuring template rows: {e}")
    
    def _map_rbmf_data_to_overview(self, rbmf_final_df: pd.DataFrame, df_output: pd.DataFrame):
        """Map Project ID and Strategic Outcome from RBMF data to Overview tab fields.
        
        Args:
            rbmf_final_df: DataFrame with RBMF_Final data
            df_output: Output DataFrame for Overview tab
        """
        try:
            if rbmf_final_df.empty:
                logger.warning("RBMF_Final data is empty, cannot map to Overview")
                return
            
            # Extract Project ID from first row
            if 'Project ID' in rbmf_final_df.columns:
                project_id = rbmf_final_df.iloc[0]['Project ID']
                if pd.notna(project_id) and str(project_id).strip():
                    # Find Project ID field in Overview tab and populate the cell to its right
                    # Based on 2025 template structure, Project ID is typically around row 2-3
                    for row_idx in range(len(df_output)):
                        for col_idx in range(len(df_output.columns)):
                            cell_value = df_output.iloc[row_idx, col_idx]
                            if pd.notna(cell_value) and 'Project ID' in str(cell_value):
                                # Fill the cell to the right
                                if col_idx + 1 < len(df_output.columns):
                                    df_output.iloc[row_idx, col_idx + 1] = str(project_id)
                                    logger.info(f"Mapped Project ID '{project_id}' to Overview tab")
                                break

                    # Try to map Implementation period start/end from project_dates.json using Project ID
                    try:
                        project_dates = self._load_project_dates()
                        if project_dates and str(project_id) in project_dates:
                            start_date, end_date = project_dates[str(project_id)]
                            # Only write when dates are non-empty and not '-'
                            if isinstance(start_date, str) and start_date.strip() and start_date.strip() != '-':
                                # Locate the "Implementation period" label and set adjacent cells
                                for r in range(len(df_output)):
                                    for c in range(len(df_output.columns)):
                                        v = df_output.iloc[r, c]
                                        if pd.notna(v) and 'Implementation period' in str(v):
                                            if c + 1 < len(df_output.columns):
                                                df_output.iloc[r, c + 1] = start_date.strip()
                                            if isinstance(end_date, str) and end_date.strip() and end_date.strip() != '-' and c + 2 < len(df_output.columns):
                                                df_output.iloc[r, c + 2] = end_date.strip()
                                            logger.info("Mapped Implementation period from project_dates.json")
                                            raise StopIteration
                        else:
                            logger.info("No matching entry in project_dates.json for Project ID; leaving Implementation period unchanged")
                    except StopIteration:
                        pass
                    except Exception as e:
                        logger.warning(f"Failed to map Implementation period from project_dates.json: {e}")
            
            # Extract Strategic Outcome from first row
            if 'Strategic Outcome' in rbmf_final_df.columns:
                strategic_outcome = rbmf_final_df.iloc[0]['Strategic Outcome']
                if pd.notna(strategic_outcome) and str(strategic_outcome).strip():
                    # Find Primary strategic outcome field in Overview tab and replace its value
                    for row_idx in range(len(df_output)):
                        for col_idx in range(len(df_output.columns)):
                            cell_value = df_output.iloc[row_idx, col_idx]
                            if pd.notna(cell_value) and 'Primary strategic outcome' in str(cell_value):
                                # Replace the cell to the right
                                if col_idx + 1 < len(df_output.columns):
                                    df_output.iloc[row_idx, col_idx + 1] = str(strategic_outcome)
                                    logger.info(f"Mapped Strategic Outcome '{strategic_outcome}' to Overview tab")
                                break
            
            logger.info("✓ RBMF data mapped to Overview tab successfully")
            
        except Exception as e:
            logger.warning(f"Failed to map RBMF data to Overview: {e}")

    def _load_project_dates(self) -> dict:
        """Load project start/end dates from project_dates.json into a mapping.
        
        Returns:
            Dict[str, Tuple[str, str]] mapping Project ID -> (start_date, end_date)
        """
        try:
            import json
            from pathlib import Path
            # Resolve repository root (assuming this file is at /app/src/rbmf_processor/...)
            repo_root = Path(__file__).resolve().parents[2]
            # Primary path
            json_path = repo_root / 'project_dates.json'
            if not json_path.exists():
                # Fallback: try data directory if provided differently
                alt_path = repo_root / 'data' / 'project_dates.json'
                json_path = alt_path if alt_path.exists() else json_path
            if not json_path.exists():
                logger.warning("project_dates.json not found; skipping Implementation period mapping")
                return {}
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            mapping = {}
            for entry in data:
                pid = str(entry.get('project_id', '')).strip()
                start = entry.get('start_date', '')
                end = entry.get('end_date', '')
                if pid:
                    mapping[pid] = (start, end)
            return mapping
        except Exception as e:
            logger.warning(f"Error loading project_dates.json: {e}")
            return {}
    
    def _apply_rbmf_filtering(self, rbmf_final_df: pd.DataFrame) -> pd.DataFrame:
        """Apply filtering logic to RBMF_Final data based on Strategic Outcome + Indicator name groups.
        
        New logic:
        1) Group data by Strategic Outcome + Indicator name
        2) For each group, check if any rows have Periodical Result > 0, keep them unchanged
        3) For rows with Periodical Result == 0, set Periodical Target empty
        4) If all rows have Periodical Result = 0, keep the last row (most recent) unchanged, 
           set Periodical Target empty for other rows
        
        Args:
            rbmf_final_df: DataFrame with RBMF_Final data
            
        Returns:
            Modified DataFrame with Periodical Target cleared for zero-result rows
        """
        try:
            if rbmf_final_df.empty:
                logger.warning("RBMF_Final data is empty, cannot apply filtering")
                return rbmf_final_df
            
            # Check required columns exist
            required_columns = ['Strategic Outcome', 'Indicator name', 'Periodical Result', 'Target Reporting Cycle']
            missing_columns = [col for col in required_columns if col not in rbmf_final_df.columns]
            if missing_columns:
                logger.warning(f"Missing required columns for filtering: {missing_columns}")
                return rbmf_final_df
            
            # Check if Periodical Target column exists
            if 'Periodical Target' not in rbmf_final_df.columns:
                logger.warning("Periodical Target column not found, cannot apply new filtering logic")
                return rbmf_final_df
            
            logger.info("Starting RBMF filtering process (new logic: clear Periodical Target for zero results)")
            
            # Create a copy to avoid modifying original data
            result_df = rbmf_final_df.copy()
            
            # Group by Strategic Outcome + Indicator name
            grouped = result_df.groupby(['Strategic Outcome', 'Indicator name'])
            
            for (strategic_outcome, indicator_name), group_df in grouped:
                logger.debug(f"Processing group: {strategic_outcome} + {indicator_name} ({len(group_df)} rows)")
                
                # Convert Periodical Result to numeric, handling any non-numeric values
                group_df = group_df.copy()
                group_df['Periodical Result'] = pd.to_numeric(group_df['Periodical Result'], errors='coerce')
                
                # Check if any row has Periodical Result > 0
                rows_with_results = group_df[group_df['Periodical Result'] > 0]
                
                if not rows_with_results.empty:
                    # Step 2: Keep rows with Periodical Result > 0 unchanged
                    logger.debug(f"  → Group has {len(rows_with_results)} rows with Periodical Result > 0, keeping unchanged")
                    
                    # Step 3: Set Periodical Target empty for rows with Periodical Result == 0
                    zero_result_indices = group_df[group_df['Periodical Result'] == 0].index
                    if not zero_result_indices.empty:
                        result_df.loc[zero_result_indices, 'Periodical Target'] = None
                        logger.debug(f"  → Cleared Periodical Target for {len(zero_result_indices)} rows with Periodical Result = 0")
                    
                    # Ensure positive result rows are kept unchanged (double-check)
                    positive_result_indices = group_df[group_df['Periodical Result'] > 0].index
                    logger.debug(f"  → Verified {len(positive_result_indices)} rows with Periodical Result > 0 are kept unchanged")
                else:
                    # Step 4: All rows have Periodical Result = 0
                    logger.debug(f"  → All {len(group_df)} rows have Periodical Result = 0")
                    
                    # Find the most recent Target Reporting Cycle
                    most_recent_row = self._find_most_recent_reporting_cycle(group_df)
                    if most_recent_row is not None:
                        most_recent_index = most_recent_row.index[0]
                        other_indices = group_df.index[group_df.index != most_recent_index]
                        
                        # Keep the most recent row completely unchanged (including Periodical Target)
                        logger.debug(f"  → Keeping most recent row completely unchanged: {group_df.loc[most_recent_index, 'Target Reporting Cycle']} (Periodical Target preserved)")
                        
                        # Set Periodical Target empty for other rows only
                        if not other_indices.empty:
                            result_df.loc[other_indices, 'Periodical Target'] = None
                            logger.debug(f"  → Cleared Periodical Target for {len(other_indices)} other rows (most recent row kept unchanged)")
                    else:
                        logger.warning(f"  → No valid row found for group: {strategic_outcome} + {indicator_name}")
            
            # Ensure stable order by source appearance: use original index if present
            if '__source_row_index__' in result_df.columns:
                result_df = result_df.sort_values(by='__source_row_index__', kind='stable').reset_index(drop=True)
            
            logger.info(f"✓ RBMF filtering completed: {len(rbmf_final_df)} rows processed (Periodical Target cleared for zero results)")
            return result_df
                
        except Exception as e:
            logger.error(f"Error applying RBMF filtering: {e}")
            return rbmf_final_df

    def _find_most_recent_reporting_cycle(self, group_df: pd.DataFrame) -> pd.DataFrame:
        """Find the row with the most recent Target Reporting Cycle.
        
        Args:
            group_df: DataFrame for a single group
            
        Returns:
            DataFrame with the most recent row, or None if no valid row found
        """
        try:
            # Parse Target Reporting Cycle to extract year and half
            def parse_reporting_cycle(cycle_str):
                if pd.isna(cycle_str) or not isinstance(cycle_str, str):
                    return (0, 0)  # Invalid cycle
                
                cycle_str = str(cycle_str).strip()
                # Expected format: "2025H2", "2025H1", "2024H2", etc.
                if len(cycle_str) >= 6 and cycle_str[-2:] in ['H1', 'H2']:
                    try:
                        year = int(cycle_str[:-2])
                        half = int(cycle_str[-1])  # 1 or 2
                        return (year, half)
                    except ValueError:
                        return (0, 0)
                return (0, 0)
            
            # Add parsed cycle info
            group_df = group_df.copy()
            group_df['_parsed_cycle'] = group_df['Target Reporting Cycle'].apply(parse_reporting_cycle)
            group_df['_year'] = group_df['_parsed_cycle'].apply(lambda x: x[0])
            group_df['_half'] = group_df['_parsed_cycle'].apply(lambda x: x[1])
            
            # Filter out invalid cycles
            valid_rows = group_df[(group_df['_year'] > 0) & (group_df['_half'] > 0)]
            
            if valid_rows.empty:
                logger.warning("No valid Target Reporting Cycle found in group")
                return None
            
            # Sort by year (descending) then by half (descending) to get most recent first
            sorted_rows = valid_rows.sort_values(['_year', '_half'], ascending=[False, False])
            most_recent = sorted_rows.iloc[[0]]  # Keep as DataFrame
            
            # Remove temporary columns
            most_recent = most_recent.drop(columns=['_parsed_cycle', '_year', '_half'])
            
            return most_recent
            
        except Exception as e:
            logger.error(f"Error finding most recent reporting cycle: {e}")
            return None
    
    def _copy_worksheet_formatting(self, source_ws, target_ws):
        """Copy formatting from source worksheet to target worksheet."""
        try:
            from copy import copy
            
            # Copy default sheet format (affects unspecified columns/rows)
            try:
                if hasattr(source_ws, 'sheet_format') and source_ws.sheet_format is not None:
                    if getattr(source_ws.sheet_format, 'defaultColWidth', None) is not None:
                        target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
                    if getattr(source_ws.sheet_format, 'defaultRowHeight', None) is not None:
                        target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
            except Exception:
                pass

            # Copy column widths and hidden flags
            for col_idx, column_dimension in source_ws.column_dimensions.items():
                target_ws.column_dimensions[col_idx].width = column_dimension.width
                try:
                    target_ws.column_dimensions[col_idx].hidden = column_dimension.hidden
                    target_ws.column_dimensions[col_idx].outlineLevel = column_dimension.outlineLevel
                except Exception:
                    pass


            # Ensure every visible column up to template max_column has an explicit width
            try:
                from openpyxl.utils import get_column_letter
                default_col_width = getattr(source_ws.sheet_format, 'defaultColWidth', None)
                if not default_col_width:
                    default_col_width = 8.43  # Excel default
                for c in range(1, source_ws.max_column + 1):
                    letter = get_column_letter(c)
                    # Skip A-H as they're set above
                    if letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                        src_w = source_ws.column_dimensions.get(letter).width if letter in source_ws.column_dimensions else None
                        tgt_dim = target_ws.column_dimensions[letter]
                        tgt_dim.width = src_w if src_w is not None else default_col_width
            except Exception:
                pass
            
            # Copy row heights and hidden flags
            for row_idx, row_dimension in source_ws.row_dimensions.items():
                target_ws.row_dimensions[row_idx].height = row_dimension.height
                try:
                    target_ws.row_dimensions[row_idx].hidden = row_dimension.hidden
                    target_ws.row_dimensions[row_idx].outlineLevel = row_dimension.outlineLevel
                except Exception:
                    pass

            # Ensure every row up to template max_row has an explicit height
            try:
                default_row_height = getattr(source_ws.sheet_format, 'defaultRowHeight', None)
                if not default_row_height:
                    default_row_height = 15
                for r in range(1, source_ws.max_row + 1):
                    src_h = source_ws.row_dimensions.get(r).height if r in source_ws.row_dimensions else None
                    target_ws.row_dimensions[r].height = src_h if src_h is not None else default_row_height
            except Exception:
                pass
            
            # Set auto-fit row height for all rows in Overview sheet
            try:
                for row_idx in range(1, target_ws.max_row + 1):
                    target_ws.row_dimensions[row_idx].height = None  # Auto-fit height
            except Exception as e:
                logger.warning(f"Failed to set auto-fit row height: {e}")
                pass
            
            # Fix text wrapping for merged cell F2:H2 (dropdown options)
            try:
                from openpyxl.styles import Alignment
                cell_f2 = target_ws['F2']
                # Create new alignment object to avoid immutable style error
                new_alignment = Alignment(
                    horizontal=cell_f2.alignment.horizontal if cell_f2.alignment else None,
                    vertical=cell_f2.alignment.vertical if cell_f2.alignment else None,
                    text_rotation=cell_f2.alignment.text_rotation if cell_f2.alignment else None,
                    wrap_text=False,  # Disable text wrapping to prevent line breaks
                    shrink_to_fit=cell_f2.alignment.shrink_to_fit if cell_f2.alignment else None,
                    indent=cell_f2.alignment.indent if cell_f2.alignment else None
                )
                cell_f2.alignment = new_alignment
                logger.info("✓ Disabled text wrapping for F2:H2 dropdown cell")
            except Exception as e:
                logger.warning(f"Failed to fix text wrapping for F2:H2: {e}")
                pass
            
            # Copy cell formatting for all cells (including empty ones)
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws.cell(row=cell.row, column=cell.column)
                    
                    # Copy formatting properties
                    if cell.font:
                        target_cell.font = copy(cell.font)
                    if cell.border:
                        target_cell.border = copy(cell.border)
                    if cell.fill:
                        target_cell.fill = copy(cell.fill)
                    if cell.number_format:
                        target_cell.number_format = cell.number_format
                    if cell.protection:
                        target_cell.protection = copy(cell.protection)
                    if cell.alignment:
                        target_cell.alignment = copy(cell.alignment)
            
            # Copy merged cells
            for merged_range in source_ws.merged_cells.ranges:
                target_ws.merge_cells(str(merged_range))
            
            # Set specific column widths for Overview tab (A-H) AFTER merged cells
            # A=181px, B=127px, C=127px, D=100px, E=222px, F=300px, G=300px, H=300px
            try:
                from openpyxl.utils import get_column_letter
                # Convert Google Sheets pixels to Excel units: ColumnWidth = 1 + (Pixels - 12) / 7
                def pixels_to_excel_units(pixels):
                    return 1 + (pixels - 12) / 7
                
                width_map = {
                    'A': pixels_to_excel_units(181),  # ~25.1
                    'B': pixels_to_excel_units(127),  # ~17.4
                    'C': pixels_to_excel_units(127),  # ~17.4
                    'D': pixels_to_excel_units(100),  # ~13.6
                    'E': pixels_to_excel_units(222),  # ~31.0
                    'F': pixels_to_excel_units(300),  # ~42.1 (restored to original)
                    'G': pixels_to_excel_units(300),  # ~42.1 (restored to original)
                    'H': pixels_to_excel_units(400),  # ~56.4 (kept wide for F2:H2 dropdown)
                }
                for col_letter, width in width_map.items():
                    target_ws.column_dimensions[col_letter].width = width
            except Exception as e:
                logger.warning(f"Failed to set specific column widths: {e}")
                pass

            # Copy freeze panes and basic sheet view options
            try:
                target_ws.freeze_panes = source_ws.freeze_panes
                if hasattr(source_ws, 'sheet_view') and hasattr(target_ws, 'sheet_view'):
                    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
                    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
                    target_ws.sheet_view.zoomScaleNormal = source_ws.sheet_view.zoomScaleNormal
            except Exception:
                pass

            # Copy print/page setup options
            try:
                target_ws.print_options = copy(source_ws.print_options)
                target_ws.page_margins = copy(source_ws.page_margins)
                target_ws.page_setup = copy(source_ws.page_setup)
            except Exception:
                pass

            # Copy sheet properties (e.g., tab color)
            try:
                if hasattr(source_ws, 'sheet_properties') and hasattr(target_ws, 'sheet_properties'):
                    target_ws.sheet_properties.tabColor = getattr(source_ws.sheet_properties, 'tabColor', None)
            except Exception:
                pass

            # Copy data validations
            try:
                from copy import copy as _copy
                if getattr(source_ws, 'data_validations', None):
                    # Recreate DataValidations container on target and append copies
                    target_ws.data_validations = type(source_ws.data_validations)()
                    for dv in source_ws.data_validations.dataValidation:
                        # Fix F2 dropdown formula to handle commas in option text
                        if dv.sqref and 'F2' in str(dv.sqref) and dv.type == 'list' and dv.formula1:
                            # Parse the original formula and fix comma handling
                            original_formula = dv.formula1
                            if original_formula.startswith('"') and original_formula.endswith('"'):
                                # Remove outer quotes
                                content = original_formula[1:-1]
                                
                                # Define the correct options
                                correct_options = [
                                    "SO1 - Policy alignment with climate commitments",
                                    "SO2 - De-risking Investments on RE, EE and Fossil Fuel Phasedown",
                                    "SO3 - Sustainable and Resilient Infrastructure",
                                    "SO4 - Just Transition"
                                ]
                                
                                # Create properly formatted formula with quoted options
                                fixed_formula = ','.join(f'"{opt}"' for opt in correct_options)
                                
                                # Create a copy of the data validation with fixed formula
                                fixed_dv = _copy(dv)
                                fixed_dv.formula1 = fixed_formula
                                target_ws.data_validations.append(fixed_dv)
                                
                                logger.info(f"✓ Fixed F2 dropdown formula to handle commas properly")
                            else:
                                # Use original formula if not in expected format
                                target_ws.data_validations.append(_copy(dv))
                        else:
                            # Copy other data validations as-is
                            target_ws.data_validations.append(_copy(dv))
            except Exception as e:
                logger.warning(f"Failed to copy data validations: {e}")
                pass
                
            logger.info("✓ Worksheet formatting copied successfully")
        except Exception as e:
            logger.warning(f"Failed to copy worksheet formatting: {e}")

    def _set_wrap_text_for_all_cells(self, worksheet) -> None:
        """Set wrap_text=True for all cells in the given worksheet while preserving alignment."""
        try:
            from openpyxl.styles import Alignment
            for row in worksheet.iter_rows():
                for cell in row:
                    existing = cell.alignment
                    cell.alignment = Alignment(
                        horizontal=getattr(existing, 'horizontal', None),
                        vertical=getattr(existing, 'vertical', None),
                        textRotation=getattr(existing, 'textRotation', 0),
                        wrap_text=True,
                        shrinkToFit=getattr(existing, 'shrinkToFit', None),
                        indent=getattr(existing, 'indent', 0)
                    )
            logger.info("✓ Wrap text enabled for all cells in Overview")
        except Exception as e:
            logger.warning(f"Failed to set wrap text on cells: {e}")

    def _copy_worksheet_values(self, source_ws, target_ws) -> None:
        """Deprecated: previously used to clone template Overview. Kept for backward compatibility."""
        return
    
    def _auto_adjust_row_heights(self, worksheet):
        """Auto-adjust row heights to fit content using improved AutoFit logic.
        
        Args:
            worksheet: The worksheet to adjust row heights for
        """
        try:
            logger.info("Starting auto row height adjustment for Overview tab")
            
            # Calculate row heights based on content with better logic
            # Only process rows that likely contain content (first 50 rows should be enough)
            max_content_row = min(50, worksheet.max_row)
            for row_idx in range(1, max_content_row + 1):
                # Respect hidden rows copied from template
                row_dim = worksheet.row_dimensions.get(row_idx)
                if row_dim is not None and getattr(row_dim, 'hidden', False):
                    continue
                max_height = 15  # Default minimum height
                max_text_length = 0
                
                # Find the longest text in this row
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        text_length = len(str(cell.value).strip())
                        max_text_length = max(max_text_length, text_length)
                
                if max_text_length > 0:
                    # Calculate height based on text length and font size
                    # Assuming average font size of 11pt and ~8 characters per inch
                    # Excel's default row height is 15 points for 11pt font
                    
                    # Estimate lines needed (more conservative estimate)
                    chars_per_line = 60  # Conservative estimate for typical column width
                    estimated_lines = max(1, (max_text_length // chars_per_line) + 1)
                    
                    # Calculate height: base height + additional height for extra lines
                    # Each line needs about 15 points of height
                    calculated_height = estimated_lines * 15
                    
                    # Add some padding for better readability
                    calculated_height += 5
                    
                    # Cap maximum height to prevent extremely tall rows
                    calculated_height = min(calculated_height, 120)
                    
                    max_height = max(max_height, calculated_height)
                
                # Apply the calculated height to the row
                worksheet.row_dimensions[row_idx].height = max_height
                
                # Log significant height adjustments only
                if max_height > 20 or max_text_length > 50:
                    logger.info(f"Row {row_idx} height set to {max_height:.1f} points (text length: {max_text_length})")
            
            logger.info("✓ Row heights auto-adjusted successfully")
            
        except Exception as e:
            logger.warning(f"Failed to auto-adjust row heights: {e}")
    
    def _auto_adjust_column_widths(self, worksheet):
        """Auto-adjust column widths to fit content using improved AutoFit logic.
        
        Args:
            worksheet: The worksheet to adjust column widths for
        """
        try:
            logger.info("Starting auto column width adjustment for Overview tab")
            
            # Calculate column widths based on content with better logic
            for col_idx in range(1, worksheet.max_column + 1):
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                # Respect hidden columns copied from template
                col_dim = worksheet.column_dimensions.get(column_letter)
                if col_dim is not None and getattr(col_dim, 'hidden', False):
                    continue
                max_width = 8.43  # Default minimum width (Excel default)
                max_text_length = 0
                
                # Find the longest text in this column
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        text_length = len(str(cell.value).strip())
                        max_text_length = max(max_text_length, text_length)
                
                if max_text_length > 0:
                    # Calculate width based on text length
                    # Excel's character width calculation: ~0.15 units per character for 11pt font
                    # Add padding for better readability
                    calculated_width = max(8.43, (max_text_length * 0.15) + 3)
                    
                    # Cap maximum width to prevent extremely wide columns
                    calculated_width = min(calculated_width, 60)
                    
                    max_width = max(max_width, calculated_width)
                
                # Apply the calculated width to the column
                worksheet.column_dimensions[column_letter].width = max_width
                
                # Log significant width adjustments only
                if max_width > 12 or max_text_length > 50:
                    logger.info(f"Column {column_letter} width set to {max_width:.1f} (text length: {max_text_length})")
            
            logger.info("✓ Column widths auto-adjusted successfully")
            
        except Exception as e:
            logger.warning(f"Failed to auto-adjust column widths: {e}")
    
    def _find_stakeholder_table(self, worksheet) -> tuple:
        """Find the table with Stakeholder Name, Department/Organisation Name, and Position columns.
        
        Args:
            worksheet: The worksheet to search
            
        Returns:
            tuple: (min_row, min_col, max_row, max_col) of the table, or None if not found
        """
        try:
            # Search for the header row with the three specific columns
            header_row = None
            stakeholder_col = None
            dept_col = None
            position_col = None
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_value = cell.value.strip()
                        if "Stakeholder Name" in cell_value:
                            stakeholder_col = cell.column
                            header_row = cell.row
                        elif "Department/ Organisation Name" in cell_value or "Department/Organisation Name" in cell_value:
                            dept_col = cell.column
                        elif "Position" in cell_value:
                            position_col = cell.column
                
                # If we found all three columns in the same row, we have our header
                if stakeholder_col and dept_col and position_col and header_row:
                    break
            
            if not all([stakeholder_col, dept_col, position_col, header_row]):
                logger.warning("Could not find stakeholder table headers in worksheet")
                return None
            
            # Determine table boundaries
            table_start_row = header_row
            table_start_col = min(stakeholder_col, dept_col, position_col)
            table_end_col = max(stakeholder_col, dept_col, position_col)
            
            # Find the bottom row of the table
            table_end_row = table_start_row
            for row_num in range(table_start_row + 1, table_start_row + 50):  # Search up to 50 rows
                has_data = False
                for col_num in range(table_start_col, table_end_col + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value and str(cell.value).strip():
                        has_data = True
                        break
                
                if has_data:
                    table_end_row = row_num
                else:
                    # If we find an empty row, check if the next few rows also have no data
                    # to determine if we've reached the end of the table
                    empty_rows = 0
                    for check_row in range(row_num, min(row_num + 3, table_start_row + 50)):
                        check_has_data = False
                        for col_num in range(table_start_col, table_end_col + 1):
                            check_cell = worksheet.cell(row=check_row, column=col_num)
                            if check_cell.value and str(check_cell.value).strip():
                                check_has_data = True
                                break
                        if not check_has_data:
                            empty_rows += 1
                    
                    if empty_rows >= 2:  # If we have 2+ consecutive empty rows, we've reached the end
                        break
            
            table_range = (table_start_row, table_start_col, table_end_row, table_end_col)
            logger.info(f"Found stakeholder table at range: {table_range}")
            return table_range
            
        except Exception as e:
            logger.error(f"Error finding stakeholder table: {e}")
            return None
    

    def _optimize_formatting_for_google_sheets(self, worksheet) -> None:
        """Optimize worksheet formatting for Google Sheets compatibility.
        
        Args:
            worksheet: openpyxl worksheet object to optimize
        """
        try:
            # Simplify complex formatting that might cause issues in Google Sheets
            for row_num, row in enumerate(worksheet.iter_rows(), 1):
                for col_num, cell in enumerate(row, 1):
                    if cell.has_style:
                        # Preserve font size 8 and specific text wrapping for RBMF tab
                        if worksheet.title == 'RBMF':
                            # For RBMF tab, preserve our specific formatting
                            if cell.font:
                                # Keep existing font but ensure size is 8
                                cell.font = openpyxl.styles.Font(
                                    name=cell.font.name,
                                    size=8,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    vertAlign=cell.font.vertAlign,
                                    underline=cell.font.underline,
                                    strike=cell.font.strike,
                                    color=cell.font.color
                                )
                            
                            # Preserve specific text wrapping for columns C, D, I
                            if cell.alignment:
                                wrap_text = col_num in [3, 4, 9]  # Columns C, D, I
                                cell.alignment = openpyxl.styles.Alignment(
                                    vertical=cell.alignment.vertical,
                                    horizontal=cell.alignment.horizontal,
                                    wrap_text=wrap_text,
                                    text_rotation=cell.alignment.text_rotation,
                                    shrink_to_fit=cell.alignment.shrink_to_fit,
                                    indent=cell.alignment.indent
                                )
                        else:
                            # For other tabs, use simplified formatting
                            # Simplify font formatting
                            if cell.font and cell.font.bold:
                                cell.font = openpyxl.styles.Font(bold=True, size=8)
                            else:
                                cell.font = openpyxl.styles.Font(size=8)
                            
                            # Simplify alignment for other tabs
                            if cell.alignment:
                                cell.alignment = openpyxl.styles.Alignment(
                                    vertical='center',
                                    horizontal='left',
                                    wrap_text=True
                                )
                        
                        # Simplify fill formatting (use basic colors only)
                        if cell.fill and cell.fill.fill_type == 'solid':
                            # Convert to basic color if it's a complex pattern
                            if hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color=cell.fill.start_color,
                                    end_color=cell.fill.start_color,
                                    fill_type="solid"
                                )
                        
                        # Simplify border formatting
                        if cell.border:
                            # Use basic border style
                            cell.border = openpyxl.styles.Border(
                                left=openpyxl.styles.Side(style='thin'),
                                right=openpyxl.styles.Side(style='thin'),
                                top=openpyxl.styles.Side(style='thin'),
                                bottom=openpyxl.styles.Side(style='thin')
                            )
            
            logger.debug("Optimized formatting for Google Sheets compatibility")
            
        except Exception as e:
            logger.warning(f"Error optimizing formatting for Google Sheets: {e}")
    
    def _write_dataframe_to_worksheet(self, df: pd.DataFrame, worksheet, header_style: Optional[Dict[str, Any]] = None) -> None:
        """Write DataFrame to Excel worksheet with optional header style.
        
        Args:
            df: DataFrame to write
            worksheet: openpyxl worksheet object
            header_style: Optional dict of style attributes to apply to header row
        """
        # Set header row height for better readability
        worksheet.row_dimensions[1].height = 25
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=column_title)
            if header_style:
                if header_style.get('font') is not None:
                    cell.font = header_style['font']
                if header_style.get('fill') is not None:
                    cell.fill = header_style['fill']
                if header_style.get('border') is not None:
                    cell.border = header_style['border']
                if header_style.get('alignment') is not None:
                    cell.alignment = header_style['alignment']
                if header_style.get('number_format') is not None:
                    cell.number_format = header_style['number_format']
            else:
                # Fallback generic header style
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data with improved row heights
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            # Set row height for better readability
            worksheet.row_dimensions[row_num].height = 20
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                
                # Set alignment for better readability
                cell.alignment = openpyxl.styles.Alignment(
                    vertical='center',
                    horizontal='left',
                    wrap_text=True
                )
        
        # Auto-adjust column widths with minimum width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ensure minimum width for readability
            min_width = 15
            adjusted_width = max(min_width, min(max_length + 2, 50))  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _write_dataframe_to_worksheet_with_template_formatting(self, df: pd.DataFrame, worksheet, template_wb) -> None:
        """Write DataFrame to Excel worksheet with template RBMF formatting.
        
        Args:
            df: DataFrame to write
            worksheet: openpyxl worksheet object
            template_wb: Template workbook to copy formatting from
        """
        try:
            # Remove the internal __source_row_index__ column from final output
            df = df.copy()
            if '__source_row_index__' in df.columns:
                df = df.drop(columns=['__source_row_index__'])
            
            # Get template RBMF worksheet for formatting reference
            if 'RBMF' in template_wb.sheetnames:
                template_rbmf_ws = template_wb['RBMF']
                
                # Write headers with exact template formatting
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num, value=column_title)
                    
                    # Copy exact formatting from template RBMF header (first row)
                    if col_num <= template_rbmf_ws.max_column:
                        template_cell = template_rbmf_ws.cell(row=1, column=col_num)
                        
                        # Copy font with exact template properties
                        if template_cell.font:
                            cell.font = openpyxl.styles.Font(
                                name=template_cell.font.name,
                                size=template_cell.font.size,
                                bold=template_cell.font.bold,
                                italic=template_cell.font.italic,
                                color=template_cell.font.color
                            )
                        
                        # Copy fill with exact template properties
                        if template_cell.fill and template_cell.fill.start_color:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=template_cell.fill.start_color,
                                end_color=template_cell.fill.end_color,
                                fill_type=template_cell.fill.fill_type
                            )
                        
                        # Copy border with exact template properties
                        if template_cell.border:
                            cell.border = openpyxl.styles.Border(
                                left=template_cell.border.left,
                                right=template_cell.border.right,
                                top=template_cell.border.top,
                                bottom=template_cell.border.bottom
                            )
                        
                        # Copy alignment with exact template properties
                        if template_cell.alignment:
                            cell.alignment = openpyxl.styles.Alignment(
                                horizontal=template_cell.alignment.horizontal,
                                vertical=template_cell.alignment.vertical,
                                wrap_text=template_cell.alignment.wrap_text,
                                text_rotation=template_cell.alignment.text_rotation,
                                shrink_to_fit=template_cell.alignment.shrink_to_fit,
                                indent=template_cell.alignment.indent
                            )
                        
                        # Copy number format
                        if template_cell.number_format:
                            cell.number_format = template_cell.number_format
                    else:
                        # Fallback formatting for extra columns
                        cell.font = openpyxl.styles.Font(bold=True, size=8)
                        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        cell.alignment = openpyxl.styles.Alignment(
                            vertical='center',
                            horizontal='center',
                            wrap_text=True
                        )
                
                # Copy exact column widths from template
                for col_num in range(1, min(template_rbmf_ws.max_column + 1, len(df.columns) + 1)):
                    col_letter = openpyxl.utils.get_column_letter(col_num)
                    template_width = template_rbmf_ws.column_dimensions[col_letter].width
                    if template_width:
                        worksheet.column_dimensions[col_letter].width = template_width
                
                # Copy exact row heights from template
                template_row_height = template_rbmf_ws.row_dimensions[1].height
                if template_row_height:
                    worksheet.row_dimensions[1].height = template_row_height
                
                # Write data with template formatting and exact row heights
                for row_num, row_data in enumerate(df.itertuples(index=False), 2):
                    # Copy row height from template if available
                    if row_num <= template_rbmf_ws.max_row:
                        template_row_height = template_rbmf_ws.row_dimensions[row_num].height
                        if template_row_height:
                            worksheet.row_dimensions[row_num].height = template_row_height
                    
                    for col_num, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=row_num, column=col_num, value=value)
                        
                        # Copy data cell formatting from template if available
                        if row_num <= template_rbmf_ws.max_row and col_num <= template_rbmf_ws.max_column:
                            template_cell = template_rbmf_ws.cell(row=row_num, column=col_num)
                            
                            # Copy font with exact template properties
                            if template_cell.font:
                                cell.font = openpyxl.styles.Font(
                                    name=template_cell.font.name,
                                    size=template_cell.font.size,
                                    bold=template_cell.font.bold,
                                    italic=template_cell.font.italic,
                                    color=template_cell.font.color
                                )
                            
                            # Copy fill with exact template properties
                            if template_cell.fill and template_cell.fill.start_color:
                                cell.fill = openpyxl.styles.PatternFill(
                                    start_color=template_cell.fill.start_color,
                                    end_color=template_cell.fill.end_color,
                                    fill_type=template_cell.fill.fill_type
                                )
                            
                            # Copy border with exact template properties
                            if template_cell.border:
                                cell.border = openpyxl.styles.Border(
                                    left=template_cell.border.left,
                                    right=template_cell.border.right,
                                    top=template_cell.border.top,
                                    bottom=template_cell.border.bottom
                                )
                            
                            # Copy alignment with exact template properties
                            if template_cell.alignment:
                                cell.alignment = openpyxl.styles.Alignment(
                                    horizontal=template_cell.alignment.horizontal,
                                    vertical=template_cell.alignment.vertical,
                                    wrap_text=template_cell.alignment.wrap_text,
                                    text_rotation=template_cell.alignment.text_rotation,
                                    shrink_to_fit=template_cell.alignment.shrink_to_fit,
                                    indent=template_cell.alignment.indent
                                )
                            
                            # Copy number format
                            if template_cell.number_format:
                                cell.number_format = template_cell.number_format
                        
                        # Ensure font size is 8 for all cells (fallback for cells without template formatting)
                        if cell.font is None:
                            cell.font = openpyxl.styles.Font(size=8)
                        
                        # Set text wrapping for specific columns (C=3, D=4, I=9)
                        wrap_text = col_num in [3, 4, 9]
                        
                        # Enhance alignment for better readability
                        if cell.alignment is None:
                            cell.alignment = openpyxl.styles.Alignment(
                                vertical='center',
                                horizontal='left',
                                wrap_text=wrap_text
                            )
                        else:
                            # Create new alignment object to avoid immutable style error
                            cell.alignment = openpyxl.styles.Alignment(
                                vertical='center',
                                horizontal='left',
                                wrap_text=wrap_text,
                                text_rotation=cell.alignment.text_rotation,
                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                indent=cell.alignment.indent
                            )
                
                # Column widths are already set from template above
                
                logger.info("Applied template RBMF formatting to RBMF_Final tab")
            else:
                # Fallback to standard formatting if template RBMF not found
                logger.warning("Template RBMF sheet not found, using standard formatting")
                self._write_dataframe_to_worksheet(df, worksheet)
                
        except Exception as e:
            logger.warning(f"Error applying template formatting: {e}, using standard formatting")
            self._write_dataframe_to_worksheet(df, worksheet)
    
    def _is_excel_file(self, file_path: Path) -> bool:
        """Check if a file is an Excel file by content.
        
        Args:
            file_path: Path to the file
            
        Returns:
            True if the file is an Excel file, False otherwise
        """
        try:
            # Try to read as Excel file
            pd.read_excel(file_path, nrows=1)
            return True
        except:
            try:
                # Try with openpyxl
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                wb.close()
                return True
            except:
                return False
    
    def create_instructions_files_for_all_folders(self) -> Dict[str, Any]:
        """Create Instructions-only files for all source files in target folders.
        
        Returns:
            Dictionary with creation results
        """
        logger.info("Starting Instructions file creation process")
        
        # Load template instructions
        self.load_template_instructions()
        
        # Create output structure
        self.create_output_structure()
        
        # Process each target folder
        results = {
            'total_files': 0,
            'created_files': 0,
            'failed_files': 0,
            'folder_results': {}
        }
        
        for folder_name in self.target_folders:
            source_folder = self.data_dir / folder_name
            output_folder = self.output_dir / folder_name
            
            if not source_folder.exists():
                logger.warning(f"Source folder does not exist: {source_folder}")
                continue
            
            logger.info(f"Processing folder: {folder_name}")
            
            folder_results = {
                'files_created': 0,
                'files_failed': 0,
                'file_results': []
            }
            
            # Process all files in the folder (including those without extensions)
            for file_path in source_folder.glob('*'):
                if file_path.is_file():
                    # Check if it's an Excel file by content or extension
                    if (file_path.suffix.lower() in ['.xlsx', '.xls'] or 
                        self._is_excel_file(file_path)):
                        results['total_files'] += 1
                        
                        # Create output file with same name
                        output_file = output_folder / file_path.name
                        
                        # Create Instructions-only file
                        success = self.create_instructions_only_file(output_file)
                        
                        file_result = {
                            'file_name': file_path.name,
                            'created': success,
                            'output_file': str(output_file) if success else None,
                            'error': None if success else "Failed to create file"
                        }
                        
                        folder_results['file_results'].append(file_result)
                        
                        if success:
                            results['created_files'] += 1
                            folder_results['files_created'] += 1
                        else:
                            results['failed_files'] += 1
                            folder_results['files_failed'] += 1
            
            results['folder_results'][folder_name] = folder_results
        
        logger.info(f"Instructions file creation completed. Created: {results['created_files']}, Failed: {results['failed_files']}")
        return results