"""Main entry point for RBMF Data Processor."""

import json
from pathlib import Path
from loguru import logger
import click

from .rbmf_processor.config import settings
from .rbmf_processor.data_processor import DataProcessor
from .rbmf_processor.gdown_client import GDownClient
from .rbmf_processor.rbmf_transformer import RBMFTransformer
from .rbmf_processor.optimized_transformer import OptimizedRBMFTransformer


def _interactive_folder_selection(data_dir: Path) -> list:
    """Show interactive menu for folder selection.
    
    Args:
        data_dir: Path to the data directory
        
    Returns:
        List of selected folder names
    """
    # Create a temporary transformer to discover folders
    temp_transformer = RBMFTransformer(data_dir)
    available_folders = temp_transformer.discover_available_folders()
    
    if not available_folders:
        print("❌ No folders with Excel files found in the data directory.")
        return []
    
    print("\n📁 Available folders in data/:")
    print("=" * 50)
    
    for i, folder in enumerate(available_folders, 1):
        print(f"{i:2d}. {folder}")
    
    print(f"{len(available_folders) + 1:2d}. All folders")
    print("=" * 50)
    
    while True:
        try:
            selection = input(f"\nSelect folders (comma-separated numbers, 1-{len(available_folders) + 1}): ").strip()
            
            if not selection:
                print("❌ No selection made. Please try again.")
                continue
            
            # Parse selection
            selected_indices = []
            for part in selection.split(','):
                part = part.strip()
                if '-' in part:
                    # Handle ranges like "1-3"
                    start, end = map(int, part.split('-'))
                    selected_indices.extend(range(start, end + 1))
                else:
                    selected_indices.append(int(part))
            
            # Validate indices
            max_index = len(available_folders) + 1
            invalid_indices = [i for i in selected_indices if i < 1 or i > max_index]
            
            if invalid_indices:
                print(f"❌ Invalid selection: {invalid_indices}. Please use numbers between 1 and {max_index}")
                continue
            
            # Convert to folder names
            selected_folders = []
            for index in selected_indices:
                if index == max_index:  # "All folders" option
                    selected_folders = available_folders.copy()
                    break
                else:
                    selected_folders.append(available_folders[index - 1])
            
            # Remove duplicates while preserving order
            selected_folders = list(dict.fromkeys(selected_folders))
            
            print(f"\n✅ Selected folders: {selected_folders}")
            return selected_folders
            
        except ValueError:
            print("❌ Invalid input. Please enter numbers separated by commas.")
        except KeyboardInterrupt:
            print("\n❌ Operation cancelled by user.")
            return []


def setup_logging():
    """Setup logging configuration."""
    logger.remove()  # Remove default handler
    
    # Add console handler (only if not in quiet mode)
    if not settings.quiet_mode:
        logger.add(
            sink=lambda msg: print(msg, end=""),
            level=settings.log_level,
            format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <8}</level> | <cyan>{name}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan> - <level>{message}</level>"
        )
    else:
        # Quiet mode: only show WARNING and ERROR on console
        logger.add(
            sink=lambda msg: print(msg, end=""),
            level="WARNING",
            format="<red>{time:HH:mm:ss}</red> | <level>{level: <8}</level> | <level>{message}</level>"
        )
    
    # Add file handler (always enabled for debugging)
    log_level = settings.log_level if settings.verbose_logging else "WARNING"
    logger.add(
        sink=settings.log_dir / "rbmf_processor.log",
        level=log_level,
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <8} | {name}:{function}:{line} - {message}",
        rotation="10 MB",
        retention="7 days"
    )


@click.group()
def cli():
    """RBMF Data Processor - Download and process Google Drive data."""
    pass


@cli.command()
@click.option('--folder-url', 
              default=settings.google_drive_folder_url,
              help='Google Drive folder URL to download')
@click.option('--output-dir', default=settings.data_dir, 
              help='Output directory for downloaded files')
def download(folder_url: str, output_dir: Path):
    """Download files from Google Drive folder using gdown (no API key required)."""
    logger.info(f"Starting download from: {folder_url}")
    
    try:
        # Initialize gdown client
        gdown_client = GDownClient()
        
        # Download folder
        downloaded_files = gdown_client.download_folder(folder_url, Path(output_dir))
        
        logger.info(f"Download completed. {len(downloaded_files)} files downloaded")
        
        # Save download summary
        summary = {
            'folder_url': folder_url,
            'download_path': str(output_dir),
            'files_downloaded': len(downloaded_files),
            'downloaded_files': [str(f) for f in downloaded_files]
        }
        
        summary_file = Path(output_dir) / "download_summary.json"
        with open(summary_file, 'w') as f:
            json.dump(summary, f, indent=2)
        
        logger.info(f"Download summary saved to: {summary_file}")
        
    except Exception as e:
        logger.error(f"Download failed: {e}")
        raise


@cli.command()
@click.option('--data-dir', default=settings.data_dir, 
              help='Directory containing data files to process')
@click.option('--output-file', default=None, 
              help='Output file for processing report (JSON)')
def process(data_dir: Path, output_file: str):
    """Process downloaded data files."""
    logger.info(f"Starting data processing from: {data_dir}")
    
    try:
        # Initialize data processor
        processor = DataProcessor()
        
        # Process all files in directory
        results = processor.process_directory(Path(data_dir))
        
        if not results:
            logger.warning("No files found to process")
            return
        
        # Generate summary report
        summary = processor.generate_summary_report(results)
        
        logger.info(f"Processing completed. {summary['successful_files']}/{summary['total_files']} files processed successfully")
        
        # Save report
        if output_file:
            report_path = Path(output_file)
        else:
            report_path = Path(data_dir) / "processing_report.json"
        
        with open(report_path, 'w') as f:
            json.dump(summary, f, indent=2, default=str)
        
        logger.info(f"Processing report saved to: {report_path}")
        
    except Exception as e:
        logger.error(f"Processing failed: {e}")
        raise


@cli.command()
@click.option('--data-dir', default=settings.data_dir, 
              help='Directory containing downloaded data files')
@click.option('--folders', multiple=True, 
              help='Specific folders to process (can be used multiple times)')
@click.option('--interactive', is_flag=True, 
              help='Show interactive folder selection menu')
@click.option('--steps', is_flag=True, 
              help='Create intermediate tabs (RBMF_1, RBMF_2) for verification')
@click.option('--filter', is_flag=True, 
              help='Apply filtering to RBMF tab based on Strategic Outcome + Indicator name groups')
@click.option('--report-file', default='transformation_report.json', 
              help='Output report file name')
@click.option('--quiet', is_flag=True, 
              help='Quiet mode - minimal console output for better performance')
def transform(data_dir: Path, folders: tuple, interactive: bool, steps: bool, filter: bool, report_file: str, quiet: bool):
    """Transform RBMF data from quarterly to half-yearly format.
    
    Default: Creates Instructions + RBMF tabs only (efficient for production)
    --steps: Creates Instructions + RBMF_1 + RBMF_2 + RBMF tabs (for verification)
    --folders: Process specific folders (can be used multiple times)
    --interactive: Show interactive folder selection menu
    --filter: Apply filtering to RBMF tab based on Strategic Outcome + Indicator name groups
    --quiet: Minimal console output for better performance
    """
    # Enable quiet mode if requested
    if quiet:
        settings.quiet_mode = True
        settings.verbose_logging = False
    
    # Setup logging after quiet mode is configured
    setup_logging()
    
    try:
        if steps:
            mode_desc = "with intermediate steps"
        elif filter:
            mode_desc = "with filtering applied"
        else:
            mode_desc = "final only"
        logger.info(f"Starting RBMF data transformation ({mode_desc}) from: {data_dir}")
        
        # Determine target folders
        target_folders = None
        
        if interactive:
            # Interactive mode - show menu and get user selection
            target_folders = _interactive_folder_selection(Path(data_dir))
            if not target_folders:
                logger.info("No folders selected. Exiting.")
                return
        elif folders:
            # Specific folders provided via command line
            target_folders = list(folders)
            logger.info(f"Processing specified folders: {target_folders}")
        else:
            # Default behavior - process all available folders
            logger.info("Processing all available folders")
        
        # Initialize transformer with steps flag and target folders
        transformer = RBMFTransformer(Path(data_dir), include_steps=steps, target_folders=target_folders)
        
        # Validate folders if specific folders were provided
        if target_folders:
            valid_folders, invalid_folders = transformer.validate_folders(target_folders)
            if invalid_folders:
                logger.error(f"Cannot proceed with invalid folders: {invalid_folders}")
                return
            # Update transformer with validated folders
            transformer.target_folders = valid_folders
        
        # Load template instructions
        transformer.load_template_instructions()
        
        # Create output structure
        if steps:
            output_mode = "steps"
        elif filter:
            output_mode = "final-filter"
        else:
            output_mode = "final"
        
        transformer.create_output_structure(output_mode)
        
        # Process each target folder
        # Determine mode description
        if steps:
            mode_desc = 'steps'
        elif filter:
            mode_desc = 'final-filter'
        else:
            mode_desc = 'final'
        
        results = {
            'total_files': 0,
            'created_files': 0,
            'failed_files': 0,
            'mode': mode_desc,
            'folder_results': {}
        }
        
        for folder_name in transformer.target_folders:
            source_folder = transformer.data_dir / folder_name
            
            # Determine output folder name
            if steps:
                output_folder_name = "steps"
            elif filter:
                output_folder_name = "final-filter"
            else:
                output_folder_name = "final"
            
            output_folder = transformer.output_dir / folder_name / output_folder_name
            
            if not source_folder.exists():
                logger.warning(f"Source folder does not exist: {source_folder}")
                continue
            
            logger.info(f"Processing folder: {folder_name}")
            
            folder_results = {
                'files_created': 0,
                'files_failed': 0,
                'file_results': []
            }
            
            # Get all files to process
            files_to_process = list(source_folder.glob('*'))
            files_to_process = [f for f in files_to_process if f.is_file() and transformer._is_excel_file(f)]
            
            # Check if parallel processing should be used
            use_parallel = (
                settings.parallel_processing and 
                len(files_to_process) > 1
            )
            
            if use_parallel:
                logger.info(f"Using parallel processing for {len(files_to_process)} files in folder: {folder_name}")
                
                # Use optimized transformer for parallel processing
                optimized_transformer = OptimizedRBMFTransformer(
                    data_dir=data_dir,
                    include_steps=steps,
                    target_folders=[folder_name],
                    apply_filter=filter
                )
                
                # Process folder with parallel processing
                parallel_results = optimized_transformer._process_folder_optimized(folder_name)
                
                # Convert parallel results to match expected format
                folder_results['files_created'] = parallel_results['files_created']
                folder_results['files_failed'] = parallel_results['files_failed']
                folder_results['file_results'] = []
                
                for file_result in parallel_results['file_results']:
                    # Convert parallel result format to expected format
                    converted_result = {
                        'file_name': file_result['file_name'],
                        'created': file_result['success'],
                        'output_file': file_result.get('output_file'),
                        'error': file_result.get('error')
                    }
                    folder_results['file_results'].append(converted_result)
                
                # Update global results
                results['total_files'] += parallel_results['files_processed']
                results['created_files'] += parallel_results['files_created']
                results['failed_files'] += parallel_results['files_failed']
                
            else:
                logger.info(f"Using sequential processing for {len(files_to_process)} files in folder: {folder_name}")
                
                # Process all files in the folder sequentially
                for file_path in files_to_process:
                    results['total_files'] += 1
                    
                    # Create output file with same name
                    output_file = output_folder / file_path.name
                    
                    # Create output file
                    success = transformer.create_output_file(file_path, output_file, apply_filter=filter)
                    
                    # Final step: apply template Overview formatting (do NOT change values)
                    if success:
                        try:
                            # Load template workbook once per iteration (cheap enough) to ensure freshness
                            template_wb = transformer.load_template_workbook()
                            if 'Overview' in template_wb.sheetnames:
                                import openpyxl
                                res_wb = openpyxl.load_workbook(output_file)
                                if 'Overview' in res_wb.sheetnames:
                                    tmpl_ws = template_wb['Overview']
                                    res_ws = res_wb['Overview']
                                    # Unmerge existing merges to avoid duplicate/conflicting ranges
                                    try:
                                        existing_merges = list(res_ws.merged_cells.ranges)
                                        for rng in existing_merges:
                                            res_ws.unmerge_cells(str(rng))
                                    except Exception:
                                        pass
                                    # Copy full formatting/layout from template (values unchanged)
                                    transformer._copy_worksheet_formatting(tmpl_ws, res_ws)
                                    # Save the updated workbook
                                    res_wb.save(output_file)
                                    logger.info(f"Applied Overview formatting from template to: {output_file}")
                                else:
                                    logger.warning(f"Output file missing 'Overview' sheet, skipped formatting: {output_file}")
                            else:
                                logger.warning("Template missing 'Overview' sheet, skipped formatting step")
                        except Exception as fmt_err:
                            logger.warning(f"Failed to apply Overview formatting to {output_file}: {fmt_err}")
                    
                    file_result = {
                        'file_name': file_path.name,
                        'created': success,
                        'output_file': str(output_file) if success else None,
                        'error': None if success else "Failed to create file"
                    }
                    
                    folder_results['file_results'].append(file_result)
                    
                    if success:
                        results['created_files'] += 1
                        folder_results['files_created'] += 1
                    else:
                        results['failed_files'] += 1
                        folder_results['files_failed'] += 1
            
            results['folder_results'][folder_name] = folder_results
        
        # Save results to report file
        report_path = Path(data_dir) / report_file
        with open(report_path, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        
        logger.info(f"Transformation completed ({mode_desc}). Report saved to: {report_path}")
        logger.info(f"Total files processed: {results['total_files']}")
        logger.info(f"Successfully created: {results['created_files']}")
        logger.info(f"Failed: {results['failed_files']}")
        
    except Exception as e:
        logger.error(f"Error during transformation: {e}")
        raise


@cli.command()
@click.option('--data-dir', default=settings.data_dir, 
              help='Data directory containing project folders')
@click.option('--output-file', default='project_names.json',
              help='Output JSON file name')
def list_projects(data_dir: Path, output_file: str):
    """Extract and list all project names from the 5 main folders.
    
    Extracts file names from: 1 INO, 2 PHI, 3 VIE, 4 REG, 5 Retainers
    Saves project names to JSON format in 2025-output folder.
    """
    try:
        logger.info(f"Starting project name extraction from: {data_dir}")
        
        # Define the 5 target folders
        target_folders = ['1 INO', '2 PHI', '3 VIE', '4 REG', '5 Retainers']
        
        # Initialize transformer to use its folder discovery methods
        transformer = RBMFTransformer(Path(data_dir))
        
        # Check if folders exist
        existing_folders = []
        missing_folders = []
        
        for folder_name in target_folders:
            folder_path = Path(data_dir) / folder_name
            if folder_path.exists() and folder_path.is_dir():
                existing_folders.append(folder_name)
            else:
                missing_folders.append(folder_name)
        
        if missing_folders:
            logger.warning(f"Missing folders: {missing_folders}")
        
        if not existing_folders:
            logger.error("No target folders found. Cannot proceed.")
            return
        
        # Extract project names from each folder
        project_data = {
            'extraction_date': str(Path().cwd()),
            'total_projects': 0,
            'folders': {}
        }
        
        for folder_name in existing_folders:
            folder_path = Path(data_dir) / folder_name
            projects = []
            
            # Get all files in the folder
            for file_path in folder_path.glob('*'):
                if file_path.is_file():
                    # Skip hidden files and system files
                    if not file_path.name.startswith('.') and not file_path.name.startswith('~'):
                        projects.append(file_path.name)
            
            # Sort projects alphabetically
            projects.sort()
            
            project_data['folders'][folder_name] = {
                'folder_name': folder_name,
                'project_count': len(projects),
                'projects': projects
            }
            project_data['total_projects'] += len(projects)
            
            logger.info(f"Found {len(projects)} projects in {folder_name}")
        
        # Create 2025-output directory if it doesn't exist
        output_dir = Path(data_dir) / '2025-output'
        output_dir.mkdir(exist_ok=True)
        
        # Save to JSON file
        output_path = output_dir / output_file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(project_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Project names saved to: {output_path}")
        logger.info(f"Total projects found: {project_data['total_projects']}")
        logger.info(f"Folders processed: {len(existing_folders)}")
        
        # Print summary
        print(f"\n📋 Project Names Summary:")
        print(f"Total projects: {project_data['total_projects']}")
        print(f"Output file: {output_path}")
        print(f"\nProjects by folder:")
        for folder_name, folder_data in project_data['folders'].items():
            print(f"  {folder_name}: {folder_data['project_count']} projects")
        
    except Exception as e:
        logger.error(f"Error during project name extraction: {e}")
        raise


@cli.command()
@click.option('--folder-url', 
              default=settings.google_drive_folder_url,
              help='Google Drive folder URL to download')
@click.option('--output-dir', default=settings.data_dir, 
              help='Output directory for downloaded files')
def run(folder_url: str, output_dir: Path):
    """Download and transform data in one command."""
    logger.info("Starting full pipeline: download + transform")
    
    # First download
    ctx = click.get_current_context()
    ctx.invoke(download, folder_url=folder_url, output_dir=output_dir)
    
    # Then transform
    ctx.invoke(transform, data_dir=output_dir)


@cli.command()
@click.option('--template-file', required=True, help='Path to the template Excel file (with Overview sheet)')
@click.option('--result-file', required=True, help='Path to the result Excel file whose Overview data must retain, but adopt template formatting')
def format_overview(template_file: str, result_file: str):
    """Apply full formatting/layout from template's Overview to result's Overview without changing data."""
    from openpyxl import load_workbook
    from pathlib import Path as _Path
    
    setup_logging()
    logger.info("Starting Overview formatting application (template -> result)")
    try:
        template_path = _Path(template_file)
        result_path = _Path(result_file)
        if not template_path.exists():
            logger.error(f"Template file not found: {template_path}")
            return
        if not result_path.exists():
            logger.error(f"Result file not found: {result_path}")
            return
        
        # Load workbooks
        tmpl_wb = load_workbook(template_path)
        res_wb = load_workbook(result_path)
        
        if 'Overview' not in tmpl_wb.sheetnames:
            logger.error("Template workbook missing 'Overview' sheet")
            return
        if 'Overview' not in res_wb.sheetnames:
            logger.error("Result workbook missing 'Overview' sheet")
            return
        
        tmpl_ws = tmpl_wb['Overview']
        res_ws = res_wb['Overview']
        
        # Unmerge any existing merges in result to avoid duplicate/conflicting merge ranges
        try:
            existing_merges = list(res_ws.merged_cells.ranges)
            for rng in existing_merges:
                res_ws.unmerge_cells(str(rng))
        except Exception:
            pass
        
        # Reuse transformer formatting copier (does not overwrite values)
        transformer = RBMFTransformer(data_dir=_Path('.'))
        transformer._copy_worksheet_formatting(tmpl_ws, res_ws)
        
        # Preserve the data values (already unchanged); save workbook
        res_wb.save(result_path)
        logger.info(f"✓ Applied template formatting to Overview without altering data: {result_path}")
        
    except Exception as e:
        logger.error(f"Failed to apply Overview formatting: {e}")


if __name__ == "__main__":
    cli()
